import asyncio
import logging
import sys
import sqlite3
import json
import random
import openpyxl
import io
import os
import re
import hashlib
import time
from collections import defaultdict
from pathlib import Path
from datetime import datetime, timedelta

BASE_DIR = Path(__file__).resolve().parent

from aiogram import Bot, Dispatcher, F, Router
from aiogram.filters import CommandStart
from aiogram.types import (
    Message,
    ReplyKeyboardMarkup,
    KeyboardButton,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    CallbackQuery,
    BufferedInputFile,
    InputMediaPhoto,
    ChatMemberUpdated,
    ChatPermissions
)
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.enums import ParseMode
from aiogram.filters import CommandStart
from aiogram.filters.command import CommandObject
from aiogram.filters import ChatMemberUpdatedFilter, JOIN_TRANSITION
from aiogram.exceptions import TelegramBadRequest, TelegramForbiddenError

# -----------------------------------------------------------------------------
# NASTROYKALAR
# -----------------------------------------------------------------------------
TOKEN = "8550058562:AAFMl5EK-i1REFGGQIsIPyCGBYNwPpTh9dQ"
ADMIN_IDS = [2110945697, 1924892484]
STOP_NARKO_BOT_USERNAME = "stopnarko_bot"
BOT_USERNAME = "FrontOfisBot"

# ── MAJBURIY KANAL ───────────────────────────────────────────────────────────
# @ belgisi BILAN yozing, masalan: "@antinarko_front"
# Bo'sh qoldirsa (None) — tekshiruv o'chiriladi
REQUIRED_CHANNELS = ["@YoshlarFrontOfisi", "@BaxaTech2025"]

CERTIFICATE_THRESHOLD = 0.86
CERTIFICATE_TEMPLATE_PATH = str(BASE_DIR / "certificates" / "template.png")

# ── A'ZOLIK GUVOHNOMASI (ID-karta) ───────────────────────────────────────────
GUV_FRONT_PATH = str(BASE_DIR / "certificates" / "guvohnoma_old.png")
GUV_BACK_PATH  = str(BASE_DIR / "certificates" / "guvohnoma_orqa.png")
GUV_FONT_BOLD  = str(BASE_DIR / "certificates" / "fonts" / "Lato-Bold.ttf")
GUV_FONT_REG   = str(BASE_DIR / "certificates" / "fonts" / "Lato-Regular.ttf")

# Amal qilish muddati — shablonda chop etilgan, o'zgarmaydi
GUV_EXPIRE_DATE = "31.08.2027"

# Hudud kodlari: ID raqamining birinchi 2 ta raqami
GUV_REGIONS = [
    ("01", "Toshkent shahri"),
    ("02", "Toshkent viloyati"),
    ("03", "Andijon"),
    ("04", "Farg'ona"),
    ("05", "Namangan"),
    ("06", "Samarqand"),
    ("07", "Buxoro"),
    ("08", "Xorazm"),
    ("09", "Surxondaryo"),
    ("10", "Qashqadaryo"),
    ("11", "Jizzax"),
    ("12", "Sirdaryo"),
    ("13", "Navoiy"),
    ("14", "Qoraqalpog'iston Respublikasi"),
]
GUV_REGION_NAME = {code: name for code, name in GUV_REGIONS}

# Excel'dan kelgan turli yozilishlarni kodga bog'lash
_GUV_REGION_ALIASES = {
    "toshkent shahri": "01", "toshkent shahar": "01", "toshkent sh": "01",
    "toshkent city": "01", "tashkent city": "01", "toshkent shaxri": "01",
    "toshkent viloyati": "02", "toshkent viloyat": "02", "toshkent vil": "02",
    "andijon": "03", "andijon viloyati": "03", "andijan": "03", "andijon vil": "03",
    "fargona": "04", "fargona viloyati": "04", "fergana": "04", "farghona": "04",
    "namangan": "05", "namangan viloyati": "05",
    "samarqand": "06", "samarqand viloyati": "06", "samarkand": "06",
    "buxoro": "07", "buxoro viloyati": "07", "bukhara": "07", "buxara": "07",
    "xorazm": "08", "xorazm viloyati": "08", "khorezm": "08", "urganch": "08",
    "surxondaryo": "09", "surxondaryo viloyati": "09", "surxandaryo": "09",
    "qashqadaryo": "10", "qashqadaryo viloyati": "10", "kashkadarya": "10",
    "jizzax": "11", "jizzax viloyati": "11", "jizzakh": "11",
    "sirdaryo": "12", "sirdaryo viloyati": "12", "syrdarya": "12",
    "navoiy": "13", "navoiy viloyati": "13", "navoi": "13",
    "qoraqalpogiston": "14", "qoraqalpogiston respublikasi": "14",
    "qoraqalpogiston res": "14", "karakalpakstan": "14", "nukus": "14",
}

# =============================================================================
# MODERATSIYA SOZLAMALARI
# =============================================================================
# Ogohlantirish tizimi
WARN_LIMIT = 3                  # Nechta ogohlantirishdan keyin ban
MUTE_DAYS_PER_WARN = 7         # Har bir ogohlantirish uchun mute kunlari

# Anti-flood
RATE_LIMIT_MAX = 5             # Max xabarlar soni
RATE_LIMIT_WINDOW = 10         # Soniyalar oralig'i
RATE_LIMIT_MUTE = 120          # Flood uchun mute vaqti (soniya)

# Tekshiruvlar (True/False)
CHECK_BANNED_WORDS = True
CHECK_HACKER_WORDS = True
CHECK_LINKS = True
CHECK_SHELL_PATTERNS = True
BLOCK_DANGEROUS_FILES = True

# Taqiqlangan so'zlar ro'yxati
BANNED_WORDS = [
    # Reklama/daromad
    "заработай", "зарабатывай", "заработок", "быстрые деньги",
    "пассивный доход", "казино", "ставки", "без вложений",
    # Narkotiklar
    "наркотики", "наркота", "закладка", "закладки",
    "мефедрон", "героин", "кокаин", "амфетамин", "спайс",
    "gashish", "narkotik", "nasha", "mef", "nasva",
    # VPN reklama
    "bepul vpn", "vpn yuklab", "blokirovkani aylanish",
]

# Xakerlik so'zlari
HACKER_WORDS = [
    "kali linux", "metasploit", "bettercap", "ettercap",
    "aircrack", "burpsuite", "sqlmap", "hydra", "nmap",
    "wireshark", "hashcat", "netcat", "nikto",
    "mitm", "man-in-the-middle", "arp spoofing", "phishing",
    "keylogger", "reverse shell", "brute force", "sql injection",
    "xss attack", "ddos attack",
    "powershell -e", "powershell -enc", "invoke-expression",
    "iex(", "downloadstring", "shellcode", "msfvenom",
    "msfconsole", "cmd /c",
    "взломать", "взлом сайта", "эксплойт", "вирус скачать",
]

# Bloklanadigan fayl kengaytmalari
BLOCKED_EXTENSIONS = [
    "exe", "bat", "cmd", "ps1", "vbs", "scr", "pif",
    "com", "msi", "dll", "reg", "hta", "wsf", "lnk",
]

# Havolalar uchun pattern
LINK_PATTERN = re.compile(
    r"(https?://|t\.me/|bit\.ly/|tinyurl\.com/|vk\.com/|instagram\.com/)",
    re.IGNORECASE
)


# =============================================================================
# MAJBURIY KANAL TEKSHIRUVI
# =============================================================================
async def check_subscription(bot: Bot, user_id: int) -> bool:
    """Foydalanuvchi barcha kanallarga a'zo ekanligini tekshiradi."""
    if not REQUIRED_CHANNELS:
        return True
    for channel in REQUIRED_CHANNELS:
        try:
            member = await bot.get_chat_member(channel, user_id)
            if member.status in ["left", "kicked", "banned"]:
                return False
        except Exception as e:
            logging.warning(f"Kanal tekshirishda xato ({channel}): {e}")
    return True


def sub_required_kb() -> InlineKeyboardMarkup:
    """Har bir kanal uchun alohida tugma + tekshirish tugmasi."""
    buttons = [
        [InlineKeyboardButton(
            text=f"📢 {ch} kanaliga a'zo bo'lish",
            url=f"https://t.me/{ch.lstrip('@')}"
        )]
        for ch in REQUIRED_CHANNELS
    ]
    buttons.append(
        [InlineKeyboardButton(text="✅ A'zo bo'ldim, tekshirish", callback_data="check_sub")]
    )
    return InlineKeyboardMarkup(inline_keyboard=buttons)


# -----------------------------------------------------------------------------
# DATABASE
# -----------------------------------------------------------------------------
class Database:
    def __init__(self, db_name="yoshlar_fronti.db"):
        db_path = BASE_DIR / db_name
        self.conn = sqlite3.connect(str(db_path), check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self.cursor = self.conn.cursor()
        self.create_tables()

    def create_tables(self):
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS password_resets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                passport_data TEXT,
                full_name TEXT,
                status TEXT DEFAULT 'pending'
            )
        """)
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS incidents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                description TEXT,
                date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS tests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                description TEXT,
                time_limit INTEGER DEFAULT 30,
                question_count INTEGER DEFAULT 10,
                max_attempts INTEGER DEFAULT 0,
                is_active INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        # Migration: max_attempts ustuni qo'shish (eski DB uchun)
        self.cursor.execute("PRAGMA table_info(tests)")
        test_cols = [row[1] for row in self.cursor.fetchall()]
        if "max_attempts" not in test_cols:
            self.cursor.execute("ALTER TABLE tests ADD COLUMN max_attempts INTEGER DEFAULT 0")

        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS questions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                test_id INTEGER NOT NULL,
                question_text TEXT NOT NULL,
                option_a TEXT NOT NULL,
                option_b TEXT NOT NULL,
                option_c TEXT NOT NULL,
                option_d TEXT NOT NULL,
                correct_answer TEXT NOT NULL,
                FOREIGN KEY (test_id) REFERENCES tests(id) ON DELETE CASCADE
            )
        """)
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS test_results (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                user_name TEXT,
                test_id INTEGER NOT NULL,
                score INTEGER DEFAULT 0,
                total INTEGER DEFAULT 0,
                passed INTEGER DEFAULT 0,
                certificate_issued INTEGER DEFAULT 0,
                cert_code TEXT,
                finished_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (test_id) REFERENCES tests(id)
            )
        """)
        self.cursor.execute("PRAGMA table_info(test_results)")
        columns = [row[1] for row in self.cursor.fetchall()]
        if "cert_code" not in columns:
            self.cursor.execute("ALTER TABLE test_results ADD COLUMN cert_code TEXT")

        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY,
                full_name TEXT,
                username TEXT,
                phone TEXT,
                region TEXT,
                birth_year TEXT,
                registered_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        existing = [row[1] for row in self.cursor.execute("PRAGMA table_info(users)").fetchall()]
        for col, col_type in [("phone", "TEXT"), ("region", "TEXT"), ("birth_year", "TEXT")]:
            if col not in existing:
                self.cursor.execute(f"ALTER TABLE users ADD COLUMN {col} {col_type}")
        self.conn.commit()

    def register_user(self, user_id, username):
        self.cursor.execute(
            "INSERT OR IGNORE INTO users (id, username) VALUES (?, ?)",
            (user_id, username)
        )
        self.conn.commit()

    def get_user(self, user_id):
        self.cursor.execute("SELECT * FROM users WHERE id=?", (user_id,))
        return self.cursor.fetchone()

    def is_registered(self, user_id):
        self.cursor.execute(
            "SELECT id FROM users WHERE id=? AND full_name IS NOT NULL AND full_name != ''",
            (user_id,)
        )
        return self.cursor.fetchone() is not None

    def set_user_name(self, user_id, full_name):
        self.cursor.execute("UPDATE users SET full_name=? WHERE id=?", (full_name, user_id))
        self.conn.commit()

    def update_user_profile(self, user_id: int, full_name: str = None, phone: str = None,
                             region: str = None, birth_year: str = None):
        user = self.get_user(user_id)
        if not user:
            return
        self.cursor.execute(
            "UPDATE users SET full_name=?, phone=?, region=?, birth_year=? WHERE id=?",
            (
                full_name if full_name is not None else user["full_name"],
                phone if phone is not None else user["phone"],
                region if region is not None else user["region"],
                birth_year if birth_year is not None else user["birth_year"],
                user_id
            )
        )
        self.conn.commit()

    def get_active_tests(self):
        self.cursor.execute("SELECT * FROM tests WHERE is_active=1 ORDER BY created_at DESC")
        return self.cursor.fetchall()

    def get_all_tests(self):
        self.cursor.execute("SELECT * FROM tests ORDER BY created_at DESC")
        return self.cursor.fetchall()

    def get_test(self, test_id):
        self.cursor.execute("SELECT * FROM tests WHERE id=?", (test_id,))
        return self.cursor.fetchone()

    def create_test(self, title, description, time_limit, question_count, max_attempts=0):
        self.cursor.execute(
            "INSERT INTO tests (title, description, time_limit, question_count, max_attempts) VALUES (?, ?, ?, ?, ?)",
            (title, description, time_limit, question_count, max_attempts)
        )
        self.conn.commit()
        return self.cursor.lastrowid

    def update_test(self, test_id, title, description, time_limit, question_count, max_attempts=None):
        if max_attempts is None:
            self.cursor.execute(
                "UPDATE tests SET title=?, description=?, time_limit=?, question_count=? WHERE id=?",
                (title, description, time_limit, question_count, test_id)
            )
        else:
            self.cursor.execute(
                "UPDATE tests SET title=?, description=?, time_limit=?, question_count=?, max_attempts=? WHERE id=?",
                (title, description, time_limit, question_count, max_attempts, test_id)
            )
        self.conn.commit()

    def toggle_test(self, test_id):
        self.cursor.execute("SELECT is_active FROM tests WHERE id=?", (test_id,))
        row = self.cursor.fetchone()
        if row:
            new_status = 0 if row["is_active"] else 1
            self.cursor.execute("UPDATE tests SET is_active=? WHERE id=?", (new_status, test_id))
            self.conn.commit()
            return new_status
        return None

    def delete_test(self, test_id):
        self.cursor.execute("DELETE FROM tests WHERE id=?", (test_id,))
        self.conn.commit()

    def add_question(self, test_id, question_text, a, b, c, d, correct):
        self.cursor.execute(
            "INSERT INTO questions (test_id, question_text, option_a, option_b, option_c, option_d, correct_answer) "
            "VALUES (?, ?, ?, ?, ?, ?, ?)",
            (test_id, question_text, a, b, c, d, correct.upper())
        )
        self.conn.commit()

    def get_questions(self, test_id, limit=None):
        self.cursor.execute("SELECT * FROM questions WHERE test_id=?", (test_id,))
        rows = self.cursor.fetchall()
        if limit and len(rows) > limit:
            rows = random.sample(rows, limit)
        return rows

    def get_question_count(self, test_id):
        self.cursor.execute("SELECT COUNT(*) as cnt FROM questions WHERE test_id=?", (test_id,))
        return self.cursor.fetchone()["cnt"]

    def delete_questions(self, test_id):
        self.cursor.execute("DELETE FROM questions WHERE test_id=?", (test_id,))
        self.conn.commit()

    def get_user_test_attempt_count(self, user_id: int, test_id: int) -> int:
        """Foydalanuvchi ushbu testga necha marta uringanini qaytaradi."""
        self.cursor.execute(
            "SELECT COUNT(*) as cnt FROM test_results WHERE user_id=? AND test_id=?",
            (user_id, test_id)
        )
        return self.cursor.fetchone()["cnt"]

    def save_result(self, user_id, user_name, test_id, score, total):
        import uuid
        passed = 1 if (score / total) >= CERTIFICATE_THRESHOLD else 0
        cert_code = None
        if passed:
            raw = uuid.uuid4().hex[:8].upper()
            cert_code = f"{raw[:4]}-{raw[4:]}"
        self.cursor.execute(
            "INSERT INTO test_results (user_id, user_name, test_id, score, total, passed, certificate_issued, cert_code) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (user_id, user_name, test_id, score, total, passed, passed, cert_code)
        )
        self.conn.commit()
        return passed, cert_code

    def get_cert_by_code(self, cert_code: str):
        self.cursor.execute(
            "SELECT r.*, t.title FROM test_results r JOIN tests t ON r.test_id = t.id WHERE r.cert_code=?",
            (cert_code,)
        )
        return self.cursor.fetchone()

    def get_results(self, test_id=None, limit=None):
        """limit=None bo'lsa — barcha natijalar qaytariladi."""
        if test_id and limit:
            self.cursor.execute(
                "SELECT r.*, t.title FROM test_results r JOIN tests t ON r.test_id = t.id "
                "WHERE r.test_id=? ORDER BY r.finished_at DESC LIMIT ?", (test_id, limit)
            )
        elif test_id:
            self.cursor.execute(
                "SELECT r.*, t.title FROM test_results r JOIN tests t ON r.test_id = t.id "
                "WHERE r.test_id=? ORDER BY r.finished_at DESC", (test_id,)
            )
        elif limit:
            self.cursor.execute(
                "SELECT r.*, t.title FROM test_results r JOIN tests t ON r.test_id = t.id "
                "ORDER BY r.finished_at DESC LIMIT ?", (limit,)
            )
        else:
            self.cursor.execute(
                "SELECT r.*, t.title FROM test_results r JOIN tests t ON r.test_id = t.id "
                "ORDER BY r.finished_at DESC"
            )
        return self.cursor.fetchall()

    def get_passed_results(self):
        """Faqat sertifikat olgan natijalar."""
        self.cursor.execute(
            "SELECT r.*, t.title FROM test_results r JOIN tests t ON r.test_id = t.id "
            "WHERE r.passed=1 ORDER BY r.finished_at DESC"
        )
        return self.cursor.fetchall()

    def get_user_results(self, user_id):
        self.cursor.execute(
            "SELECT r.*, t.title FROM test_results r JOIN tests t ON r.test_id = t.id "
            "WHERE r.user_id=? ORDER BY r.finished_at DESC", (user_id,)
        )
        return self.cursor.fetchall()

    def save_reset_request(self, user_id, passport, full_name):
        self.cursor.execute(
            "INSERT INTO password_resets (user_id, passport_data, full_name) VALUES (?, ?, ?)",
            (user_id, passport, full_name)
        )
        self.conn.commit()

    def save_incident(self, user_id, description):
        self.cursor.execute(
            "INSERT INTO incidents (user_id, description) VALUES (?, ?)",
            (user_id, description)
        )
        self.conn.commit()

    def get_all_users(self):
        self.cursor.execute("SELECT * FROM users ORDER BY registered_at DESC")
        return self.cursor.fetchall()

    def get_user_count(self):
        self.cursor.execute("SELECT COUNT(*) as cnt FROM users")
        return self.cursor.fetchone()["cnt"]

    def get_kpi_stats(self):
        stats = {}
        self.cursor.execute("SELECT COUNT(*) as cnt FROM users WHERE full_name IS NOT NULL AND full_name != ''")
        stats["total_users"] = self.cursor.fetchone()["cnt"]
        self.cursor.execute("SELECT COUNT(*) as cnt FROM test_results")
        stats["total_attempts"] = self.cursor.fetchone()["cnt"]
        self.cursor.execute("SELECT COUNT(DISTINCT user_id) as cnt FROM test_results WHERE passed=1")
        stats["cert_users"] = self.cursor.fetchone()["cnt"]
        self.cursor.execute("SELECT COUNT(*) as cnt FROM test_results WHERE passed=1")
        stats["total_certs"] = self.cursor.fetchone()["cnt"]
        self.cursor.execute("SELECT COUNT(*) as cnt FROM test_results WHERE passed=0")
        stats["total_failed"] = self.cursor.fetchone()["cnt"]
        self.cursor.execute("SELECT AVG(CAST(score AS FLOAT)/total*100) as avg FROM test_results WHERE total>0")
        row = self.cursor.fetchone()
        stats["avg_score"] = round(row["avg"] or 0, 1)
        self.cursor.execute("SELECT COUNT(*) as cnt FROM test_results WHERE DATE(finished_at)=DATE('now')")
        stats["today_attempts"] = self.cursor.fetchone()["cnt"]
        self.cursor.execute("SELECT COUNT(*) as cnt FROM users WHERE DATE(registered_at)=DATE('now')")
        stats["today_users"] = self.cursor.fetchone()["cnt"]
        return stats

    def get_per_test_kpi(self):
        self.cursor.execute("""
            SELECT t.id, t.title,
                COUNT(r.id) as attempts,
                SUM(COALESCE(r.passed,0)) as passed,
                COUNT(r.id) - SUM(COALESCE(r.passed,0)) as failed,
                COUNT(DISTINCT r.user_id) as unique_users,
                ROUND(AVG(CAST(r.score AS FLOAT)/r.total*100), 1) as avg_pct
            FROM tests t
            LEFT JOIN test_results r ON t.id = r.test_id
            GROUP BY t.id ORDER BY attempts DESC
        """)
        return self.cursor.fetchall()

    def get_top_users(self, limit=10):
        self.cursor.execute("""
            SELECT r.user_id, r.user_name, COUNT(*) as certs,
                   ROUND(AVG(CAST(r.score AS FLOAT)/r.total*100), 1) as avg_pct
            FROM test_results r WHERE r.passed=1
            GROUP BY r.user_id ORDER BY certs DESC, avg_pct DESC LIMIT ?
        """, (limit,))
        return self.cursor.fetchall()

    def create_extra_tables(self):
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS dynamic_admins (
                user_id INTEGER PRIMARY KEY,
                added_by INTEGER,
                added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS broadcast_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                admin_id INTEGER,
                text TEXT,
                sent INTEGER DEFAULT 0,
                failed INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS question_answered (
                user_id INTEGER PRIMARY KEY,
                answered_by INTEGER,
                answered_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        # Front ofis a'zolari (yangi ariza topshiruvchilar)
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS front_members (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                full_name TEXT,
                age INTEGER,
                age_category TEXT,
                phone TEXT,
                region TEXT,
                district TEXT,
                workplace TEXT,
                photo_file_id TEXT,
                registered_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        # Admin yuklaydigan tasdiqlangan a'zolar ro'yxati (guvohnoma olish huquqi)
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS approved_members (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tg_id INTEGER,
                phone_key TEXT,
                phone TEXT,
                last_name TEXT,
                first_name TEXT,
                middle_name TEXT,
                region TEXT,
                position TEXT,
                added_by INTEGER,
                added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        self.cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_appr_tg ON approved_members(tg_id)")
        self.cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_appr_ph ON approved_members(phone_key)")
        # Berilgan a'zolik guvohnomalari
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS membership_cards (
                card_id TEXT PRIMARY KEY,
                user_id INTEGER,
                last_name TEXT,
                first_name TEXT,
                middle_name TEXT,
                region TEXT,
                region_code TEXT,
                position TEXT,
                phone TEXT,
                photo_file_id TEXT,
                serial TEXT,
                issued_date TEXT,
                expires_date TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        self.cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_card_user ON membership_cards(user_id)")
        self.conn.commit()

    # ── GUVOHNOMA: TASDIQLANGAN A'ZOLAR RO'YXATI ─────────────────────────
    def upsert_approved_member(self, tg_id, phone_key, phone, last_name,
                               first_name, middle_name, region, position, added_by):
        row = None
        if tg_id:
            self.cursor.execute(
                "SELECT id FROM approved_members WHERE tg_id=?", (tg_id,))
            row = self.cursor.fetchone()
        if row is None and phone_key:
            self.cursor.execute(
                "SELECT id FROM approved_members WHERE phone_key=?", (phone_key,))
            row = self.cursor.fetchone()
        if row:
            self.cursor.execute(
                """UPDATE approved_members SET tg_id=COALESCE(?, tg_id),
                       phone_key=COALESCE(?, phone_key), phone=COALESCE(?, phone),
                       last_name=COALESCE(?, last_name), first_name=COALESCE(?, first_name),
                       middle_name=COALESCE(?, middle_name), region=COALESCE(?, region),
                       position=COALESCE(?, position)
                   WHERE id=?""",
                (tg_id, phone_key, phone, last_name, first_name,
                 middle_name, region, position, row["id"])
            )
            self.conn.commit()
            return "updated"
        self.cursor.execute(
            """INSERT INTO approved_members
               (tg_id, phone_key, phone, last_name, first_name, middle_name,
                region, position, added_by)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (tg_id, phone_key, phone, last_name, first_name, middle_name,
             region, position, added_by)
        )
        self.conn.commit()
        return "added"

    def find_approved_by_tg(self, tg_id: int):
        self.cursor.execute("SELECT * FROM approved_members WHERE tg_id=?", (tg_id,))
        return self.cursor.fetchone()

    def find_approved_by_phone(self, phone_key: str):
        self.cursor.execute(
            "SELECT * FROM approved_members WHERE phone_key=?", (phone_key,))
        return self.cursor.fetchone()

    def bind_approved_tg(self, row_id: int, tg_id: int):
        self.cursor.execute(
            "UPDATE approved_members SET tg_id=? WHERE id=?", (tg_id, row_id))
        self.conn.commit()

    def get_approved_count(self) -> int:
        self.cursor.execute("SELECT COUNT(*) as cnt FROM approved_members")
        return self.cursor.fetchone()["cnt"]

    def get_all_approved(self):
        self.cursor.execute("SELECT * FROM approved_members ORDER BY id")
        return self.cursor.fetchall()

    def clear_approved(self):
        self.cursor.execute("DELETE FROM approved_members")
        self.conn.commit()

    # ── GUVOHNOMA: BERILGAN KARTALAR ─────────────────────────────────────
    def card_id_exists(self, card_id: str) -> bool:
        self.cursor.execute(
            "SELECT 1 FROM membership_cards WHERE card_id=?", (card_id,))
        return self.cursor.fetchone() is not None

    def save_membership_card(self, **kw):
        self.cursor.execute(
            """INSERT OR REPLACE INTO membership_cards
               (card_id, user_id, last_name, first_name, middle_name, region,
                region_code, position, phone, photo_file_id, serial,
                issued_date, expires_date)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (kw["card_id"], kw["user_id"], kw["last_name"], kw["first_name"],
             kw["middle_name"], kw["region"], kw["region_code"], kw["position"],
             kw.get("phone"), kw.get("photo_file_id"), kw["serial"],
             kw["issued_date"], kw["expires_date"])
        )
        self.conn.commit()

    def get_card_by_user(self, user_id: int):
        self.cursor.execute(
            "SELECT * FROM membership_cards WHERE user_id=? ORDER BY created_at DESC LIMIT 1",
            (user_id,))
        return self.cursor.fetchone()

    def get_card_by_id(self, card_id: str):
        self.cursor.execute(
            "SELECT * FROM membership_cards WHERE card_id=?", (card_id,))
        return self.cursor.fetchone()

    def get_all_cards(self):
        self.cursor.execute(
            "SELECT * FROM membership_cards ORDER BY created_at DESC")
        return self.cursor.fetchall()

    def get_card_count(self) -> int:
        self.cursor.execute("SELECT COUNT(*) as cnt FROM membership_cards")
        return self.cursor.fetchone()["cnt"]

    # ── FRONT OFIS A'ZOLARI ──────────────────────────────────────────────
    def add_front_member(self, user_id, full_name, age, age_category,
                         phone, region, district, workplace, photo_file_id):
        self.cursor.execute(
            """INSERT INTO front_members
               (user_id, full_name, age, age_category, phone, region,
                district, workplace, photo_file_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (user_id, full_name, age, age_category, phone, region,
             district, workplace, photo_file_id)
        )
        self.conn.commit()
        return self.cursor.lastrowid

    def get_all_front_members(self):
        self.cursor.execute("SELECT * FROM front_members ORDER BY registered_at DESC")
        return self.cursor.fetchall()

    def get_front_member_count(self):
        self.cursor.execute("SELECT COUNT(*) as cnt FROM front_members")
        return self.cursor.fetchone()["cnt"]

    def get_front_member_stats(self):
        stats = {}
        self.cursor.execute("SELECT COUNT(*) as cnt FROM front_members")
        stats["total"] = self.cursor.fetchone()["cnt"]
        self.cursor.execute(
            "SELECT COUNT(*) as cnt FROM front_members WHERE DATE(registered_at)=DATE('now')"
        )
        stats["today"] = self.cursor.fetchone()["cnt"]
        self.cursor.execute(
            "SELECT region, COUNT(*) as cnt FROM front_members "
            "GROUP BY region ORDER BY cnt DESC"
        )
        stats["by_region"] = self.cursor.fetchall()
        self.cursor.execute(
            "SELECT age_category, COUNT(*) as cnt FROM front_members "
            "GROUP BY age_category ORDER BY cnt DESC"
        )
        stats["by_category"] = self.cursor.fetchall()
        return stats

    def get_front_member(self, user_id: int):
        self.cursor.execute(
            "SELECT * FROM front_members WHERE user_id=? ORDER BY id DESC LIMIT 1",
            (user_id,))
        return self.cursor.fetchone()

    def is_front_member(self, user_id: int) -> bool:
        self.cursor.execute("SELECT 1 FROM front_members WHERE user_id=?", (user_id,))
        return self.cursor.fetchone() is not None

    def get_dynamic_admins(self):
        self.cursor.execute("SELECT * FROM dynamic_admins ORDER BY added_at")
        return self.cursor.fetchall()

    def add_dynamic_admin(self, user_id: int, added_by: int):
        self.cursor.execute(
            "INSERT OR IGNORE INTO dynamic_admins (user_id, added_by) VALUES (?, ?)",
            (user_id, added_by)
        )
        self.conn.commit()

    def remove_dynamic_admin(self, user_id: int):
        self.cursor.execute("DELETE FROM dynamic_admins WHERE user_id=?", (user_id,))
        self.conn.commit()

    def log_broadcast(self, admin_id: int, text: str, sent: int, failed: int):
        self.cursor.execute(
            "INSERT INTO broadcast_log (admin_id, text, sent, failed) VALUES (?, ?, ?, ?)",
            (admin_id, text, sent, failed)
        )
        self.conn.commit()

    def get_broadcast_history(self, limit=10):
        try:
            self.cursor.execute(
                "SELECT * FROM broadcast_log ORDER BY created_at DESC LIMIT ?", (limit,)
            )
            return self.cursor.fetchall()
        except Exception:
            return []

    def mark_question_answered(self, user_id: int, answered_by: int):
        self.cursor.execute(
            "INSERT OR REPLACE INTO question_answered (user_id, answered_by, answered_at) VALUES (?, ?, CURRENT_TIMESTAMP)",
            (user_id, answered_by)
        )
        self.conn.commit()

    def is_question_answered(self, user_id: int) -> bool:
        try:
            self.cursor.execute("SELECT 1 FROM question_answered WHERE user_id=?", (user_id,))
            return self.cursor.fetchone() is not None
        except Exception:
            return False

    def clear_question_answered(self, user_id: int):
        try:
            self.cursor.execute("DELETE FROM question_answered WHERE user_id=?", (user_id,))
            self.conn.commit()
        except Exception:
            pass


db = Database()
db.create_extra_tables()

SUPERADMIN_ID = ADMIN_IDS[0]


def get_all_admin_ids() -> list:
    dynamic = [row["user_id"] for row in db.get_dynamic_admins()]
    return list(set(ADMIN_IDS + dynamic))


# =============================================================================
# MODERATSIYA: IN-MEMORY SAQLASH
# =============================================================================
# { (chat_id, user_id): warn_count }
_warns: dict = defaultdict(int)
# { chat_id: { "messages": int, "last_reset": float } }
_rate_tracker: dict = defaultdict(lambda: {"messages": 0, "last_reset": time.time()})
# { notif_key: { "msgs": [...], "handled": bool } }
_notifications: dict = {}


def mod_get_warns(chat_id: int, user_id: int) -> int:
    return _warns[(chat_id, user_id)]

def mod_add_warn(chat_id: int, user_id: int) -> int:
    _warns[(chat_id, user_id)] += 1
    return _warns[(chat_id, user_id)]

def mod_reset_warns(chat_id: int, user_id: int):
    _warns[(chat_id, user_id)] = 0

def mod_save_notification(key: str, msgs: list):
    _notifications[key] = {"msgs": msgs, "handled": False}

def mod_get_notification(key: str):
    return _notifications.get(key)

def mod_mark_handled(key: str):
    if key in _notifications:
        _notifications[key]["handled"] = True


# =============================================================================
# MODERATSIYA: CONTENT FILTER FUNKSIYALARI
# =============================================================================

def filter_check_banned_word(text: str) -> str | None:
    text_lower = text.lower()
    for word in BANNED_WORDS:
        if word.lower() in text_lower:
            return word
    return None

def filter_check_hacker_word(text: str) -> str | None:
    text_lower = text.lower()
    for word in HACKER_WORDS:
        if word.lower() in text_lower:
            return word
    return None

def filter_check_link(text: str) -> str | None:
    match = LINK_PATTERN.search(text)
    return match.group(0) if match else None

def filter_check_shell(text: str) -> str | None:
    shell_patterns = [
        r'powershell\s+-\w',
        r'cmd\s+/[cskvq]',
        r'iex\s*\(',
        r'wget\s+http',
        r'curl\s+http',
        r'nc\s+-\w',
        r'python\s+-c',
        r'bash\s+-c',
        r'chmod\s+[0-7]{3}',
    ]
    for pattern in shell_patterns:
        if re.search(pattern, text, re.IGNORECASE):
            return pattern
    return None

def filter_check_file_ext(filename: str) -> str | None:
    if not filename:
        return None
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    return ext if ext in BLOCKED_EXTENSIONS else None


# =============================================================================
# MODERATSIYA: ANTI-FLOOD MIDDLEWARE
# =============================================================================
from aiogram import BaseMiddleware
from typing import Any, Awaitable, Callable

class RateLimitMiddleware(BaseMiddleware):
    """Flood xabarlarni avtomatik aniqlash va mute qilish."""

    async def __call__(
        self,
        handler: Callable[[Message, dict[str, Any]], Awaitable[Any]],
        event: Message,
        data: dict[str, Any],
    ) -> Any:
        if not hasattr(event, "from_user") or not event.from_user:
            return await handler(event, data)
        if event.chat.type == "private":
            return await handler(event, data)

        chat_id = event.chat.id
        user_id = event.from_user.id

        # Adminlarni o'tkazib yuborish
        try:
            member = await event.bot.get_chat_member(chat_id, user_id)
            if member.status in ("administrator", "creator"):
                return await handler(event, data)
        except Exception:
            pass

        now = time.time()
        tracker = _rate_tracker[(chat_id, user_id)]

        if now - tracker["last_reset"] > RATE_LIMIT_WINDOW:
            tracker["messages"] = 0
            tracker["last_reset"] = now

        tracker["messages"] += 1

        if tracker["messages"] > RATE_LIMIT_MAX:
            try:
                until = datetime.now() + timedelta(seconds=RATE_LIMIT_MUTE)
                await event.bot.restrict_chat_member(
                    chat_id=chat_id,
                    user_id=user_id,
                    permissions=ChatPermissions(can_send_messages=False),
                    until_date=until
                )
                await event.answer(
                    f"⚠️ {event.from_user.full_name}, flood uchun "
                    f"{RATE_LIMIT_MUTE // 60} daqiqaga mute qilindi!"
                )
            except Exception:
                pass
            tracker["messages"] = 0
            return  # handler ni chaqirmaymiz

        return await handler(event, data)
# -----------------------------------------------------------------------------
def is_admin(user_id: int) -> bool:
    return user_id in get_all_admin_ids()


def is_superadmin(user_id: int) -> bool:
    return user_id == SUPERADMIN_ID


async def safe_edit(message, text: str, reply_markup=None, parse_mode=ParseMode.HTML):
    try:
        await message.edit_text(text, reply_markup=reply_markup, parse_mode=parse_mode)
    except Exception:
        try:
            await message.delete()
        except Exception:
            pass
        await message.answer(text, reply_markup=reply_markup, parse_mode=parse_mode)


async def notify_admins(bot, text: str, reply_markup=None):
    for admin_id in get_all_admin_ids():
        try:
            await bot.send_message(admin_id, text, reply_markup=reply_markup, parse_mode=ParseMode.HTML)
        except Exception as e:
            logging.error(f"Admin {admin_id} ga xabar yuborilmadi: {e}")


# -----------------------------------------------------------------------------
# STATES
# -----------------------------------------------------------------------------
class RegisterState(StatesGroup):
    waiting_for_fullname = State()
    waiting_for_birth_year = State()
    waiting_for_region = State()
    waiting_for_passport = State()
    waiting_for_password = State()

class CertCheckState(StatesGroup):
    waiting_for_code = State()

class ResetPasswordState(StatesGroup):
    waiting_for_passport = State()
    waiting_for_name = State()

class QuestionState(StatesGroup):
    waiting_for_question = State()

class AdminState(StatesGroup):
    waiting_for_new_password = State()

class AdminTestState(StatesGroup):
    creating_title = State()
    creating_description = State()
    creating_time = State()
    creating_count = State()
    creating_max_attempts = State()   # YANGI
    editing_value = State()
    adding_q_text = State()
    adding_q_a = State()
    adding_q_b = State()
    adding_q_c = State()
    adding_q_d = State()
    adding_q_correct = State()
    waiting_excel = State()
    waiting_certificate = State()

class TestTakingState(StatesGroup):
    taking_test = State()
    confirming_stop = State()

class ProfileEditState(StatesGroup):
    editing_name = State()
    editing_phone = State()
    editing_region = State()
    editing_birth_year = State()

class FrontOfisState(StatesGroup):
    """Front ofis a'zoligiga ariza topshirish bosqichlari."""
    waiting_fullname = State()
    waiting_age = State()
    waiting_age_category = State()
    waiting_phone = State()
    waiting_region = State()
    waiting_district = State()
    waiting_workplace = State()
    waiting_photo = State()

class GuvState(StatesGroup):
    """A'zolik guvohnomasini olish bosqichlari."""
    waiting_phone = State()
    waiting_last_name = State()
    waiting_first_name = State()
    waiting_middle_name = State()
    waiting_region = State()
    waiting_photo = State()

class AdminGuvState(StatesGroup):
    """Admin: a'zolar ro'yxatini Excel orqali yuklash."""
    waiting_excel = State()

# -----------------------------------------------------------------------------
# KLAVIATURALAR
# -----------------------------------------------------------------------------
main_menu = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="📝 Testlar"), KeyboardButton(text="👤 Mening ma'lumotlarim")],
        [KeyboardButton(text="🏢 Front ofis a'zoligiga ariza")],
        [KeyboardButton(text="🪪 A'zolik guvohnomasi")],
        [KeyboardButton(text="🔍 Sertifikat tekshirish")],
        [KeyboardButton(text="❓ Loyiha haqida savol berish")],
        [KeyboardButton(text="🔐 Parolni tiklash"), KeyboardButton(text="🚨 Xabar berish")]
    ],
    resize_keyboard=True,
    input_field_placeholder="Biror narsa tanlang..."
)

cancel_kb = ReplyKeyboardMarkup(
    keyboard=[[KeyboardButton(text="❌ Bekor qilish")]],
    resize_keyboard=True
)

# Front ofis: telefon so'rash klaviaturasi
front_phone_kb = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="📱 Telefon raqamni yuborish", request_contact=True)],
        [KeyboardButton(text="❌ Bekor qilish")]
    ],
    resize_keyboard=True
)

# Front ofis: yosh toifalari
FRONT_AGE_CATEGORIES = [
    "14-17 yosh (o'smir)",
    "18-30 yosh (yosh)",
    "31-45 yosh (o'rta)",
    "46+ yosh (katta)",
]

# Front ofis: hududlar ro'yxati
FRONT_REGIONS = [
    "Toshkent shahri", "Toshkent viloyati", "Andijon", "Farg'ona", "Namangan",
    "Samarqand", "Buxoro", "Navoiy", "Qashqadaryo", "Surxondaryo",
    "Xorazm", "Sirdaryo", "Jizzax", "Qoraqalpog'iston"
]

router = Router()


# =============================================================================
# SERTIFIKAT GENERATSIYASI
# =============================================================================
_BOLD_FONTS = [
    "/usr/share/fonts/truetype/google-fonts/Poppins-Bold.ttf",
    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
    "/usr/share/fonts/truetype/freefont/FreeSansBold.ttf",
    "/Library/Fonts/Arial Bold.ttf",
    "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
    "/System/Library/Fonts/Helvetica.ttc",
    "C:/Windows/Fonts/arialbd.ttf",
]
_REGULAR_FONTS = [
    "/usr/share/fonts/truetype/google-fonts/Poppins-Regular.ttf",
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
    "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
    "/Library/Fonts/Arial.ttf",
    "/System/Library/Fonts/Supplemental/Arial.ttf",
    "C:/Windows/Fonts/arial.ttf",
]


def _load_font_strict(paths: list, size: int):
    from PIL import ImageFont
    for p in paths:
        if os.path.exists(p):
            try:
                return ImageFont.truetype(p, size)
            except Exception:
                pass
    mac_dirs = [
        "/System/Library/Fonts", "/System/Library/Fonts/Supplemental",
        "/Library/Fonts", os.path.expanduser("~/Library/Fonts"),
    ]
    bold_keywords = ["bold", "Bold", "Heavy", "Black", "Semibold"]
    for d in mac_dirs:
        if not os.path.isdir(d):
            continue
        for fname in os.listdir(d):
            if not (fname.endswith(".ttf") or fname.endswith(".otf") or fname.endswith(".ttc")):
                continue
            is_bold = any(k in fname for k in bold_keywords)
            if "bold" in paths[0].lower() and not is_bold:
                continue
            try:
                return ImageFont.truetype(os.path.join(d, fname), size)
            except Exception:
                pass
    try:
        import subprocess
        result = subprocess.run(["fc-list", "--format=%{file}\n"],
                                capture_output=True, text=True, timeout=3)
        for line in result.stdout.strip().split('\n'):
            line = line.strip()
            if line and (line.endswith('.ttf') or line.endswith('.otf')) and os.path.exists(line):
                try:
                    return ImageFont.truetype(line, size)
                except Exception:
                    pass
    except Exception:
        pass
    logging.warning(f"[FONT] Hech bir font topilmadi, load_default ishlatilmoqda!")
    return ImageFont.load_default()


def _find_line_y(arr, W: int, H: int) -> int:
    try:
        dark = (
            (arr[:, :, 0] < 160) &
            (arr[:, :, 1] < 160) &
            (arr[:, :, 2] < 160)
        )
        row_counts = dark.sum(axis=1)
        threshold = W * 0.30
        candidates = [(y, int(row_counts[y])) for y in range(H) if row_counts[y] > threshold]
        if not candidates:
            return int(H * 0.504)
        candidates.sort(key=lambda x: x[0])
        clusters = []
        cur_ys, cur_cnts = [candidates[0][0]], [candidates[0][1]]
        for y, cnt in candidates[1:]:
            if y - cur_ys[-1] <= 6:
                cur_ys.append(y)
                cur_cnts.append(cnt)
            else:
                clusters.append((int(sum(cur_ys) / len(cur_ys)), max(cur_cnts)))
                cur_ys, cur_cnts = [y], [cnt]
        clusters.append((int(sum(cur_ys) / len(cur_ys)), max(cur_cnts)))
        inner = [(y, c) for y, c in clusters if H * 0.15 < y < H * 0.90]
        if inner:
            best_y = max(inner, key=lambda x: x[1])[0]
        else:
            best_y = max(clusters, key=lambda x: x[1])[0]
        return best_y
    except Exception as e:
        logging.warning(f"Chiziq topishda xato: {e}")
        return int(H * 0.504)


async def generate_certificate(full_name, test_title, score, total, cert_code) -> bytes | None:
    try:
        from PIL import Image, ImageDraw, ImageFont
        import qrcode as qrcode_lib

        if not os.path.exists(CERTIFICATE_TEMPLATE_PATH):
            logging.warning("Sertifikat shabloni topilmadi: " + CERTIFICATE_TEMPLATE_PATH)
            return None

        img = Image.open(CERTIFICATE_TEMPLATE_PATH).convert("RGBA")
        W, H = img.size
        overlay = Image.new("RGBA", (W, H), (0, 0, 0, 0))
        draw = ImageDraw.Draw(overlay)
        NAME_COLOR = (14, 52, 110, 255)

        WX1 = int(W * 0.050); WX2 = int(W * 0.951); WCX = (WX1 + WX2) // 2
        LINE_Y = int(H * 0.504)
        QR_X1 = int(W * 0.449); QR_X2 = int(W * 0.562)
        QR_Y1 = int(H * 0.696); QR_Y2 = int(H * 0.860)
        QR_CX = (QR_X1 + QR_X2) // 2; QR_CY = (QR_Y1 + QR_Y2) // 2
        QR_BOX_SIZE = min(QR_X2 - QR_X1, QR_Y2 - QR_Y1)
        RIGHT_LINE_Y  = int(H * 0.791)
        RIGHT_LINE_X1 = int(W * 0.563); RIGHT_LINE_X2 = int(W * 0.847)
        RIGHT_LINE_CX = (RIGHT_LINE_X1 + RIGHT_LINE_X2) // 2

        fs_name = max(int(H * 0.085), 52); fs_code = max(int(H * 0.046), 30)
        fs_label = max(int(H * 0.027), 18); fs_hint = max(int(H * 0.023), 15)

        font_name  = _load_font_strict(_BOLD_FONTS, fs_name)
        font_code  = _load_font_strict(_BOLD_FONTS, fs_code)
        font_label = _load_font_strict(_REGULAR_FONTS, fs_label)
        font_hint  = _load_font_strict(_REGULAR_FONTS, fs_hint)

        TARGET_W = int((WX2 - WX1) * 0.72)
        while fs_name > 30:
            font_name = _load_font_strict(_BOLD_FONTS, fs_name)
            bb = draw.textbbox((0, 0), full_name, font=font_name)
            if bb[2] - bb[0] <= TARGET_W:
                break
            fs_name -= 2

        bb = draw.textbbox((0, 0), full_name, font=font_name)
        nw, nh = bb[2] - bb[0], bb[3] - bb[1]
        nx = WCX - nw // 2; ny = LINE_Y - nh - max(10, int(H * 0.008))
        draw.text((nx + 2, ny + 2), full_name, font=font_name, fill=(0, 0, 0, 40))
        draw.text((nx, ny), full_name, font=font_name, fill=NAME_COLOR)

        qr_url = f"https://t.me/{BOT_USERNAME}?start=cert_{cert_code}"
        qr_obj = qrcode_lib.QRCode(version=2, box_size=8, border=1,
                                   error_correction=qrcode_lib.constants.ERROR_CORRECT_M)
        qr_obj.add_data(qr_url); qr_obj.make(fit=True)
        qr_pil = qr_obj.make_image(fill_color="black", back_color="white").convert("RGBA")
        PAD = int(QR_BOX_SIZE * 0.04); qr_size = QR_BOX_SIZE - PAD * 2
        qr_pil = qr_pil.resize((qr_size, qr_size), Image.LANCZOS)
        qr_x = QR_CX - qr_size // 2; qr_y = QR_CY - qr_size // 2

        def cx_text(text, font, cx, y, color):
            bb = draw.textbbox((0, 0), text, font=font)
            tw, th = bb[2] - bb[0], bb[3] - bb[1]
            draw.text((cx - tw // 2, y), text, font=font, fill=color)
            return th

        bb_code  = draw.textbbox((0, 0), cert_code, font=font_code)
        bb_label = draw.textbbox((0, 0), "Sertifikat raqami", font=font_label)
        code_h = bb_code[3] - bb_code[1]; label_h = bb_label[3] - bb_label[1]
        GAP = max(6, int(H * 0.005))
        raqam_y = RIGHT_LINE_Y - code_h - GAP
        label_y = raqam_y - label_h - GAP // 2
        hint_y  = RIGHT_LINE_Y + GAP

        cx_text("Sertifikat raqami", font_label, RIGHT_LINE_CX, label_y, (90, 90, 90, 255))
        cx_text(cert_code, font_code, RIGHT_LINE_CX, raqam_y, NAME_COLOR)
        cx_text("@FrontOfisBot orqali tekshiring", font_hint, RIGHT_LINE_CX, hint_y, (100, 100, 100, 220))

        result = Image.alpha_composite(img, overlay)
        result.paste(qr_pil, (qr_x, qr_y), qr_pil)
        buf = io.BytesIO()
        result.convert("RGB").save(buf, format="PNG", dpi=(150, 150))
        logging.info(f"[CERT] OK — {len(buf.getvalue())} bytes")
        return buf.getvalue()

    except Exception as e:
        logging.error(f"[CERT] Xato: {e}", exc_info=True)
        return None


# =============================================================================
# A'ZOLIK GUVOHNOMASI (ID-KARTA) GENERATSIYASI
# =============================================================================
# Shablon o'lchami: 2024 x 1276 px (85.6 x 54 mm, 600 dpi — ID-1 standarti)
GUV_INK = (17, 17, 17)
GUV_PHOTO_BOX = (102, 364, 590, 946)      # x1, y1, x2, y2
GUV_F_FAMILIYA = (661, 1433, 515)         # x1, x2, chiziq y
GUV_F_ISM      = (675, 1447, 723)
GUV_F_OTASI    = (654, 1426, 931)
GUV_F_SANA     = (1435, 1919, 515)
GUV_F_ID       = (145, 574, 1149)
GUV_B_QR       = (266, 379, 405)          # x, y, tomon uzunligi
GUV_B_LAVOZIM  = (808, 1709, 500)
GUV_B_SERIYA   = (1443, 1893, 1197)


def guv_norm_text(s: str) -> str:
    """Turli apostroflarni bitta ko'rinishga keltiradi (oʻ, o`, o' -> o')."""
    s = (s or "").strip()
    for ch in ("ʻ", "ʼ", "`", "´", "'", "’"):
        s = s.replace(ch, "‘")
    return re.sub(r"\s+", " ", s)


def guv_phone_key(phone: str) -> str:
    """Telefon raqamdan solishtirish uchun kalit: oxirgi 9 ta raqam."""
    d = re.sub(r"\D", "", phone or "")
    return d[-9:] if len(d) >= 9 else ""


def guv_region_code(text: str) -> str | None:
    """Hudud nomidan 2 xonali kod topadi."""
    s = (text or "").lower()
    for ch in ("‘", "’", "`", "'", "ʻ", "ʼ", "´"):
        s = s.replace(ch, "")
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return None
    if s in _GUV_REGION_ALIASES:
        return _GUV_REGION_ALIASES[s]
    # eng uzun mos keluvchi alias
    best, best_len = None, 0
    for alias, code in _GUV_REGION_ALIASES.items():
        if alias in s and len(alias) > best_len:
            best, best_len = code, len(alias)
    return best


def guv_make_card_id(region_code: str) -> str:
    """Hudud kodi + 4 ta tasodifiy raqam, takrorlanmasligi tekshiriladi."""
    for _ in range(200):
        cid = f"{region_code}{random.randint(0, 9999):04d}"
        if not db.card_id_exists(cid):
            return cid
    return f"{region_code}{int(time.time()) % 10000:04d}"


def _guv_font(path: str, size: int):
    from PIL import ImageFont
    try:
        return ImageFont.truetype(path, size)
    except Exception:
        return _load_font_strict(_BOLD_FONTS if "Bold" in path else _REGULAR_FONTS, size)


def _guv_fit_font(draw, text: str, path: str, max_w: int, cap_px: int):
    """Bosh harf balandligi cap_px bo'lgan shrift; kerak bo'lsa enga sig'diradi."""
    lo, hi = 8, 400
    for _ in range(24):
        m = (lo + hi) // 2
        f = _guv_font(path, m)
        bb = f.getbbox("H")
        if (bb[3] - bb[1]) < cap_px:
            lo = m
        else:
            hi = m
    size = max(lo, 10)
    while size > 12:
        f = _guv_font(path, size)
        bb = draw.textbbox((0, 0), text, font=f)
        if (bb[2] - bb[0]) <= max_w:
            return f
        size -= 2
    return _guv_font(path, 12)


def _guv_field(draw, text: str, box, path: str, cap_px: int,
               align: str = "left", pad: int = 8):
    text = guv_norm_text(text)
    if not text:
        return
    x1, x2, y = box
    f = _guv_fit_font(draw, text, path, (x2 - x1) - 2 * pad, cap_px)
    if align == "center":
        draw.text(((x1 + x2) // 2, y), text, font=f, fill=GUV_INK, anchor="ms")
    else:
        draw.text((x1 + pad, y), text, font=f, fill=GUV_INK, anchor="ls")


def _guv_fit_photo(img, w: int, h: int):
    """3x4 rasmni ramkaga to'ldirib kesadi (bosh qismi saqlanadi)."""
    from PIL import Image
    iw, ih = img.size
    s = max(w / iw, h / ih)
    img = img.resize((max(1, int(iw * s)), max(1, int(ih * s))), Image.LANCZOS)
    left = max(0, (img.width - w) // 2)
    top = max(0, int((img.height - h) * 0.30))   # yuqoriga suriladi
    return img.crop((left, top, left + w, top + h))


async def generate_membership_card(data: dict, photo_bytes: bytes):
    """A'zolik guvohnomasining old va orqa tomonini PNG bytes ko'rinishida qaytaradi."""
    try:
        from PIL import Image, ImageDraw
        import qrcode as qrcode_lib

        for p in (GUV_FRONT_PATH, GUV_BACK_PATH):
            if not os.path.exists(p):
                logging.error(f"[GUV] Shablon topilmadi: {p}")
                return None, None

        # ── OLD TOMON ────────────────────────────────────────────────────────
        front = Image.open(GUV_FRONT_PATH).convert("RGB")
        px1, py1, px2, py2 = GUV_PHOTO_BOX
        if photo_bytes:
            try:
                ph = Image.open(io.BytesIO(photo_bytes)).convert("RGB")
                front.paste(_guv_fit_photo(ph, px2 - px1, py2 - py1), (px1, py1))
            except Exception as e:
                logging.warning(f"[GUV] Rasm joylashtirilmadi: {e}")
        d = ImageDraw.Draw(front)
        d.rectangle([px1, py1, px2 - 1, py2 - 1], outline=(206, 211, 220), width=3)

        _guv_field(d, data["last_name"], GUV_F_FAMILIYA, GUV_FONT_BOLD, 69)
        _guv_field(d, data["first_name"], GUV_F_ISM, GUV_FONT_BOLD, 69)
        _guv_field(d, data.get("middle_name", ""), GUV_F_OTASI, GUV_FONT_BOLD, 69)
        _guv_field(d, data["issued_date"], GUV_F_SANA, GUV_FONT_BOLD, 69, align="center")
        _guv_field(d, data["card_id"], GUV_F_ID, GUV_FONT_BOLD, 76, align="center")

        # ── ORQA TOMON ───────────────────────────────────────────────────────
        back = Image.open(GUV_BACK_PATH).convert("RGB")
        qx, qy, qs = GUV_B_QR
        url = f"https://t.me/{BOT_USERNAME}?start=guv_{data['card_id']}"
        qr = qrcode_lib.QRCode(version=None, box_size=10, border=0,
                               error_correction=qrcode_lib.constants.ERROR_CORRECT_Q)
        qr.add_data(url)
        qr.make(fit=True)
        qimg = qr.make_image(fill_color=GUV_INK, back_color="white").convert("RGB")
        back.paste(qimg.resize((qs, qs), Image.NEAREST), (qx, qy))

        d2 = ImageDraw.Draw(back)
        _guv_field(d2, data["position"], GUV_B_LAVOZIM, GUV_FONT_BOLD, 62)
        _guv_field(d2, data["serial"], GUV_B_SERIYA, GUV_FONT_BOLD, 76, align="center")

        out = []
        for im in (front, back):
            buf = io.BytesIO()
            im.save(buf, format="PNG", dpi=(600, 600))
            out.append(buf.getvalue())
        logging.info(f"[GUV] OK — {data['card_id']} "
                     f"({len(out[0])}+{len(out[1])} bytes)")
        return out[0], out[1]

    except Exception as e:
        logging.error(f"[GUV] Xato: {e}", exc_info=True)
        return None, None


# =============================================================================
# MAJBURIY KANAL — CALLBACK HANDLER
# =============================================================================
@router.callback_query(F.data == "check_sub")
async def check_sub_callback(callback: CallbackQuery, state: FSMContext):
    """Foydalanuvchi 'A'zo bo'ldim' tugmasini bosganida tekshiradi."""
    subscribed = await check_subscription(callback.bot, callback.from_user.id)
    if subscribed:
        await callback.answer("✅ Rahmat! Endi botdan foydalanishingiz mumkin.", show_alert=True)
        try:
            await callback.message.delete()
        except Exception:
            pass
        # Foydalanuvchini tizimga qaytarish
        user_id = callback.from_user.id
        db.register_user(user_id, callback.from_user.username or "")
        if db.is_registered(user_id):
            user = db.get_user(user_id)
            await callback.message.answer(
                f"Assalomu alaykum, <b>{user['full_name']}</b>!\n\n"
                "🏛 <b>ANTI-NARKO YOSHLAR FRONT OFISI</b> botiga xush kelibsiz.",
                reply_markup=main_menu, parse_mode=ParseMode.HTML
            )
        else:
            await state.set_state(RegisterState.waiting_for_fullname)
            await callback.message.answer(
                "🏛 <b>ANTI-NARKO YOSHLAR FRONT OFISI</b> botiga xush kelibsiz!\n\n"
                "✍️ Iltimos, <b>Ism va Familiyangizni</b> to'liq kiriting:\n"
                "<i>(masalan: Alisher Karimov)</i>",
                parse_mode=ParseMode.HTML
            )
    else:
        await callback.answer(
            "❌ Siz hali kanalga a'zo bo'lmadingiz! Iltimos, avval a'zo bo'ling.",
            show_alert=True
        )


# =============================================================================
# START — DEEP LINK BILAN
# =============================================================================
@router.message(CommandStart(deep_link=True), F.chat.type == "private")
async def command_start_deeplink(message: Message, state: FSMContext, command: CommandObject):
    payload = command.args or ""
    if payload.startswith("cert_"):
        cert_code = payload[5:].upper().strip()
        await _show_cert_info(message, cert_code)
        return
    if payload.startswith("guv_"):
        db.register_user(message.from_user.id, message.from_user.username or "")
        await _show_guv_info(message, payload[4:].strip())
        return
    await command_start_handler(message, state, command)


@router.message(CommandStart(), F.chat.type == "private")
async def command_start_handler(message: Message, state: FSMContext, command: CommandObject = None):
    user_id = message.from_user.id
    db.register_user(user_id, message.from_user.username or "")

    # ── MAJBURIY KANAL TEKSHIRUVI ────────────────────────────────────────────
    if not is_admin(user_id):
        subscribed = await check_subscription(message.bot, user_id)
        if not subscribed:
            await message.answer(
                f"👋 Assalomu alaykum!\n\n"
                f"🏛 <b>ANTI-NARKO YOSHLAR FRONT OFISI</b> botidan foydalanish uchun\n"
                f"avval rasmiy kanalimizga a'zo bo'lishingiz kerak:\n\n"
                f"{' | '.join(REQUIRED_CHANNELS)}",
                reply_markup=sub_required_kb(),
                parse_mode=ParseMode.HTML
            )
            return

    if db.is_registered(user_id):
        user = db.get_user(user_id)
        await message.answer(
            f"Assalomu alaykum, <b>{user['full_name']}</b>!\n\n"
            "🏛 <b>ANTI-NARKO YOSHLAR FRONT OFISI</b> botiga xush kelibsiz.\n\n"
            "📝 Testlar bo'limida bilimingizni sinab ko'ring va <b>sertifikat</b> oling!",
            reply_markup=main_menu, parse_mode=ParseMode.HTML
        )
    else:
        await state.set_state(RegisterState.waiting_for_fullname)
        await message.answer(
            "🏛 <b>ANTI-NARKO YOSHLAR FRONT OFISI</b> botiga xush kelibsiz!\n\n"
            "✍️ Iltimos, <b>Ism va Familiyangizni</b> to'liq kiriting:\n"
            "<i>(masalan: Alisher Karimov)</i>",
            parse_mode=ParseMode.HTML
        )


@router.message(RegisterState.waiting_for_fullname, F.text, F.chat.type == "private")
async def register_fullname(message: Message, state: FSMContext):
    full_name = message.text.strip()
    if len(full_name.split()) < 2:
        await message.answer(
            "⚠️ Iltimos, <b>Ism va Familiyangizni</b> to'liq kiriting.\n"
            "<i>(masalan: Alisher Karimov)</i>",
            parse_mode=ParseMode.HTML
        )
        return
    db.set_user_name(message.from_user.id, full_name)
    await state.clear()
    await message.answer(
        f"✅ Ro'yxatdan o'tdingiz!\n\n"
        f"Xush kelibsiz, <b>{full_name}</b>!\n\n"
        "📝 Testlar bo'limida bilimingizni sinab ko'ring va <b>sertifikat</b> oling!",
        reply_markup=main_menu, parse_mode=ParseMode.HTML
    )


@router.message(RegisterState.waiting_for_fullname, F.chat.type == "private")
async def register_fullname_invalid(message: Message):
    await message.answer(
        "✍️ Iltimos, ism-familiyangizni <b>matn</b> ko'rinishida yozing.",
        parse_mode=ParseMode.HTML
    )


# =============================================================================
# SERTIFIKAT TEKSHIRISH
# =============================================================================
async def _show_guv_info(message: Message, card_id: str):
    """QR kod / ID raqam orqali a'zolik guvohnomasining haqiqiyligini tasdiqlaydi."""
    card_id = re.sub(r"\D", "", card_id or "")
    card = db.get_card_by_id(card_id)
    if not card:
        await message.answer(
            f"❌ <b>{card_id or '—'}</b> raqamli a'zolik guvohnomasi topilmadi.\n\n"
            f"⚠️ Bu guvohnoma <b>haqiqiy emas</b> yoki raqam noto'g'ri kiritilgan.",
            reply_markup=main_menu, parse_mode=ParseMode.HTML
        )
        return
    fio = " ".join(x for x in [card["last_name"], card["first_name"],
                               card["middle_name"]] if x)
    await message.answer(
        f"✅ <b>GUVOHNOMA HAQIQIY!</b>\n"
        f"{'─' * 28}\n"
        f"👤 <b>Egasi:</b> {fio}\n"
        f"🆔 <b>ID raqami:</b> <code>{card['card_id']}</code>\n"
        f"🔖 <b>Seriya raqami:</b> <code>{card['serial']}</code>\n"
        f"🗺 <b>Hudud:</b> {card['region']}\n"
        f"💼 <b>Lavozim:</b> {card['position']}\n"
        f"📅 <b>Berilgan sana:</b> {card['issued_date']}\n"
        f"⏳ <b>Amal qilish muddati:</b> {card['expires_date']}\n\n"
        f"🏛 <i>O'zbekiston Respublikasi Yoshlar ishlari agentligi</i>",
        reply_markup=main_menu, parse_mode=ParseMode.HTML
    )


async def _show_cert_info(message: Message, cert_code: str):
    # 6 xonali raqam — a'zolik guvohnomasi ID si
    if re.fullmatch(r"\d{6}", cert_code.strip()):
        await _show_guv_info(message, cert_code.strip())
        return
    cert = db.get_cert_by_code(cert_code)
    if not cert:
        await message.answer(
            f"❌ <b>{cert_code}</b> raqamli sertifikat topilmadi.\n\nRaqamni tekshiring.",
            reply_markup=main_menu, parse_mode=ParseMode.HTML
        )
        return
    pct = int(cert["score"] / cert["total"] * 100) if cert["total"] else 0
    date_str = cert["finished_at"][:10] if cert["finished_at"] else "—"
    await message.answer(
        f"✅ <b>Sertifikat haqiqiy!</b>\n\n"
        f"👤 <b>Egasi:</b> {cert['user_name'] or '—'}\n"
        f"📋 <b>Test:</b> {cert['title']}\n"
        f"📊 <b>Natija:</b> {cert['score']}/{cert['total']} ({pct}%)\n"
        f"📅 <b>Sana:</b> {date_str}\n"
        f"🔖 <b>Raqam:</b> <code>{cert['cert_code']}</code>",
        reply_markup=main_menu, parse_mode=ParseMode.HTML
    )


# =============================================================================
# PROFIL
# =============================================================================
UZ_REGIONS = [
    "Toshkent shahri", "Toshkent viloyati", "Andijon", "Farg'ona", "Namangan",
    "Samarqand", "Buxoro", "Navoiy", "Qashqadaryo", "Surxondaryo",
    "Xorazm", "Sirdaryo", "Jizzax", "Qoraqalpog'iston"
]


def _profile_text(user) -> str:
    name = user["full_name"] or "—"; phone = user["phone"] or "—"
    region = user["region"] or "—"; birth = user["birth_year"] or "—"
    reg_date = (user["registered_at"] or "")[:10] or "—"
    tg = f"@{user['username']}" if user["username"] else "—"
    return (
        f"👤 <b>MENING MA'LUMOTLARIM</b>\n\n"
        f"📛 <b>Ism Familiya:</b> {name}\n"
        f"📱 <b>Telefon:</b> {phone}\n"
        f"🗺 <b>Viloyat:</b> {region}\n"
        f"🎂 <b>Tug'ilgan yil:</b> {birth}\n"
        f"✈️ <b>Telegram:</b> {tg}\n"
        f"📅 <b>Ro'yxatdan o'tgan:</b> {reg_date}\n"
    )


def _profile_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="✏️ Ism Familiya", callback_data="edit_profile_name"),
            InlineKeyboardButton(text="📱 Telefon", callback_data="edit_profile_phone"),
        ],
        [
            InlineKeyboardButton(text="🗺 Viloyat", callback_data="edit_profile_region"),
            InlineKeyboardButton(text="🎂 Tug'ilgan yil", callback_data="edit_profile_birth"),
        ],
        [InlineKeyboardButton(text="📊 Natijalarim", callback_data="my_results")],
    ])


@router.message(F.text == "👤 Mening ma'lumotlarim", F.chat.type == "private")
async def my_profile(message: Message, state: FSMContext):
    await state.clear()
    user = db.get_user(message.from_user.id)
    if not user:
        await message.answer("Avval /start orqali ro'yxatdan o'ting.")
        return
    await message.answer(_profile_text(user), reply_markup=_profile_kb(), parse_mode=ParseMode.HTML)


@router.callback_query(F.data == "edit_profile_name")
async def edit_name_start(callback: CallbackQuery, state: FSMContext):
    await state.set_state(ProfileEditState.editing_name)
    await callback.message.edit_text(
        "✏️ <b>Yangi Ism Familiyangizni kiriting:</b>\n\n<i>Misol: Alisher Karimov</i>",
        parse_mode=ParseMode.HTML,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ Bekor", callback_data="cancel_profile_edit")]
        ])
    )
    await callback.answer()


@router.message(ProfileEditState.editing_name, F.text, F.chat.type == "private")
async def edit_name_save(message: Message, state: FSMContext):
    name = message.text.strip()
    if len(name.split()) < 2:
        await message.answer(
            "⚠️ Iltimos, <b>Ism va Familiyani</b> to'liq kiriting.",
            parse_mode=ParseMode.HTML
        )
        return
    db.update_user_profile(message.from_user.id, full_name=name)
    await state.clear()
    user = db.get_user(message.from_user.id)
    await message.answer(f"✅ Ism yangilandi!\n\n{_profile_text(user)}", reply_markup=_profile_kb(), parse_mode=ParseMode.HTML)


@router.message(ProfileEditState.editing_name, F.chat.type == "private")
async def edit_name_invalid(message: Message):
    await message.answer("✍️ Iltimos, ismni <b>matn</b> ko'rinishida yozing.", parse_mode=ParseMode.HTML)


@router.callback_query(F.data == "edit_profile_phone")
async def edit_phone_start(callback: CallbackQuery, state: FSMContext):
    await state.set_state(ProfileEditState.editing_phone)
    kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📱 Telefon raqamni ulashish", request_contact=True)],
            [KeyboardButton(text="❌ Bekor qilish")]
        ],
        resize_keyboard=True
    )
    await callback.message.answer(
        "📱 <b>Telefon raqamingizni kiriting yoki ulashing:</b>\n\n<i>Misol: +998901234567</i>",
        parse_mode=ParseMode.HTML, reply_markup=kb
    )
    await callback.answer()


@router.message(ProfileEditState.editing_phone, F.contact, F.chat.type == "private")
async def edit_phone_contact(message: Message, state: FSMContext):
    phone = message.contact.phone_number
    if not phone.startswith("+"): phone = "+" + phone
    db.update_user_profile(message.from_user.id, phone=phone)
    await state.clear()
    user = db.get_user(message.from_user.id)
    await message.answer(f"✅ Telefon yangilandi!\n\n{_profile_text(user)}", reply_markup=_profile_kb(), parse_mode=ParseMode.HTML)


@router.message(ProfileEditState.editing_phone, F.text, F.chat.type == "private")
async def edit_phone_text(message: Message, state: FSMContext):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        user = db.get_user(message.from_user.id)
        await message.answer(_profile_text(user), reply_markup=_profile_kb(), parse_mode=ParseMode.HTML)
        return
    phone = message.text.strip()
    digits = phone.replace("+", "").replace(" ", "").replace("-", "")
    if not digits.isdigit() or len(digits) < 9:
        await message.answer("⚠️ Noto'g'ri format. Misol: +998901234567"); return
    db.update_user_profile(message.from_user.id, phone=phone)
    await state.clear()
    user = db.get_user(message.from_user.id)
    await message.answer(f"✅ Telefon yangilandi!\n\n{_profile_text(user)}", reply_markup=_profile_kb(), parse_mode=ParseMode.HTML)


@router.message(ProfileEditState.editing_phone, F.chat.type == "private")
async def edit_phone_invalid(message: Message):
    await message.answer("⚠️ Iltimos, telefon raqamni matn yoki kontakt sifatida yuboring.")


@router.callback_query(F.data == "edit_profile_region")
async def edit_region_start(callback: CallbackQuery, state: FSMContext):
    await state.set_state(ProfileEditState.editing_region)
    buttons = []
    row = []
    for i, reg in enumerate(UZ_REGIONS):
        row.append(InlineKeyboardButton(text=reg, callback_data=f"set_region_{i}"))
        if len(row) == 2:
            buttons.append(row); row = []
    if row: buttons.append(row)
    buttons.append([InlineKeyboardButton(text="❌ Bekor", callback_data="cancel_profile_edit")])
    await callback.message.edit_text("🗺 <b>Viloyatingizni tanlang:</b>", parse_mode=ParseMode.HTML,
                                     reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))
    await callback.answer()


@router.callback_query(F.data.startswith("set_region_"))
async def set_region(callback: CallbackQuery, state: FSMContext):
    idx = int(callback.data.split("_")[2])
    if idx < 0 or idx >= len(UZ_REGIONS):
        await callback.answer("Xato!"); return
    db.update_user_profile(callback.from_user.id, region=UZ_REGIONS[idx])
    await state.clear()
    user = db.get_user(callback.from_user.id)
    await safe_edit(callback.message, f"✅ Viloyat yangilandi!\n\n{_profile_text(user)}", reply_markup=_profile_kb())
    await callback.answer()


@router.callback_query(F.data == "edit_profile_birth")
async def edit_birth_start(callback: CallbackQuery, state: FSMContext):
    await state.set_state(ProfileEditState.editing_birth_year)
    current_year = datetime.now().year
    years = list(range(current_year - 15, 1979, -1))
    buttons = []; row = []
    for y in years:
        row.append(InlineKeyboardButton(text=str(y), callback_data=f"set_birth_{y}"))
        if len(row) == 5:
            buttons.append(row); row = []
    if row: buttons.append(row)
    buttons.append([InlineKeyboardButton(text="❌ Bekor", callback_data="cancel_profile_edit")])
    await callback.message.edit_text("🎂 <b>Tug'ilgan yilingizni tanlang:</b>", parse_mode=ParseMode.HTML,
                                     reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))
    await callback.answer()


@router.callback_query(F.data.startswith("set_birth_"))
async def set_birth(callback: CallbackQuery, state: FSMContext):
    year = callback.data.split("_")[2]
    db.update_user_profile(callback.from_user.id, birth_year=year)
    await state.clear()
    user = db.get_user(callback.from_user.id)
    await safe_edit(callback.message, f"✅ Tug'ilgan yil yangilandi!\n\n{_profile_text(user)}", reply_markup=_profile_kb())
    await callback.answer()


@router.callback_query(F.data == "cancel_profile_edit")
async def cancel_profile_edit(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    user = db.get_user(callback.from_user.id)
    await safe_edit(callback.message, _profile_text(user), reply_markup=_profile_kb())
    await callback.answer()


@router.message(F.text.in_({"🔍 Sertifikat tekshirish", "/check"}), F.chat.type == "private")
async def cert_check_start(message: Message, state: FSMContext):
    await state.set_state(CertCheckState.waiting_for_code)
    await message.answer(
        "🔍 <b>Sertifikat / guvohnoma tekshirish</b>\n\n"
        "• Sertifikat raqami — masalan: <code>A1B2-C3D4</code>\n"
        "• A'zolik guvohnomasi ID raqami — masalan: <code>014827</code>\n\n"
        "Raqamni kiriting:",
        reply_markup=cancel_kb, parse_mode=ParseMode.HTML
    )


@router.message(CertCheckState.waiting_for_code, F.text, F.chat.type == "private")
async def cert_check_code(message: Message, state: FSMContext):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Bekor qilindi.", reply_markup=main_menu)
        return
    await state.clear()
    await _show_cert_info(message, message.text.strip().upper())


@router.message(CertCheckState.waiting_for_code, F.chat.type == "private")
async def cert_check_code_invalid(message: Message):
    await message.answer("❗ Faqat sertifikat kodini matn ko'rinishida yuboring.")


# =============================================================================
# TESTLAR
# =============================================================================
@router.message(F.text == "📝 Testlar", F.chat.type == "private")
async def tests_menu(message: Message, state: FSMContext):
    await state.clear()
    # Kanalga a'zoligini tekshirish
    if not is_admin(message.from_user.id):
        subscribed = await check_subscription(message.bot, message.from_user.id)
        if not subscribed:
            await message.answer(
                f"📢 Testlarga kirish uchun kanalga a'zo bo'ling:",
                reply_markup=sub_required_kb()
            )
            return

    active_tests = db.get_active_tests()
    if not active_tests:
        await message.answer("📭 Hozircha faol testlar yo'q.\nYaqinda qo'shiladi!", reply_markup=main_menu)
        return

    my_results = {r["test_id"]: r for r in db.get_user_results(message.from_user.id)}
    text = "📝 <b>Mavjud testlar</b>\n\nBitta testni tanlang:\n\n"
    buttons = []

    for t in active_tests:
        result_info = ""
        if t["id"] in my_results:
            r = my_results[t["id"]]
            pct = int(r["score"] / r["total"] * 100) if r["total"] else 0
            result_info = f" ✅ {pct}%" if r["passed"] else f" ❌ {pct}%"
        # Max urinish ko'rsatish
        attempt_count = db.get_user_test_attempt_count(message.from_user.id, t["id"])
        max_att = t["max_attempts"] if t["max_attempts"] else 0
        att_info = ""
        if max_att > 0:
            att_info = f" [{attempt_count}/{max_att}]"
        buttons.append([InlineKeyboardButton(
            text=f"📋 {t['title']}{result_info}{att_info}",
            callback_data=f"start_test_{t['id']}"
        )])

    buttons.append([InlineKeyboardButton(text="📊 Mening natijalarim", callback_data="my_results")])
    buttons.append([InlineKeyboardButton(text="🏠 Asosiy menyu", callback_data="back_to_main")])
    await message.answer(text, reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons), parse_mode=ParseMode.HTML)


@router.callback_query(F.data.startswith("start_test_"))
async def show_test_info(callback: CallbackQuery, state: FSMContext):
    test_id = int(callback.data.split("_")[2])
    test = db.get_test(test_id)
    if not test:
        await callback.answer("Test topilmadi!", show_alert=True); return

    q_count = db.get_question_count(test_id)
    limit = min(test["question_count"], q_count)

    # Max urinish tekshiruvi
    max_att = test["max_attempts"] if test["max_attempts"] else 0
    attempt_count = db.get_user_test_attempt_count(callback.from_user.id, test_id)
    att_text = ""
    if max_att > 0:
        att_text = f"🔄 Urinishlar: <b>{attempt_count}/{max_att}</b>\n"
        if attempt_count >= max_att:
            await callback.message.edit_text(
                f"⛔ <b>{test['title']}</b>\n\n"
                f"Siz bu testga belgilangan maksimal urinishlar soniga ({max_att} ta) yetdingiz.\n\n"
                f"Boshqa testlarni sinab ko'ring.",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="⬅Orqaga", callback_data="back_to_tests")]
                ]),
                parse_mode=ParseMode.HTML
            )
            await callback.answer(); return

    text = (
        f"📋 <b>{test['title']}</b>\n\n"
        f"📝 {test['description'] or 'Tavsif yoq'}\n\n"
        f"❓ Savollar soni: <b>{limit}</b>\n"
        f"⏱ Vaqt: <b>{test['time_limit']} daqiqa</b>\n"
        f"🏆 Sertifikat uchun: <b>{int(CERTIFICATE_THRESHOLD * 100)}% va undan yuqori</b>\n"
        f"{att_text}\n"
        f"<i>Boshlashga tayyormisiz?</i>"
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🚀 Boshlash", callback_data=f"begin_test_{test_id}")],
        [InlineKeyboardButton(text="⬅Orqaga", callback_data="back_to_tests")]
    ])
    await callback.message.edit_text(text, reply_markup=kb, parse_mode=ParseMode.HTML)
    await callback.answer()


@router.callback_query(F.data == "back_to_tests")
async def back_to_tests(callback: CallbackQuery, state: FSMContext):
    await callback.message.delete()
    await tests_menu(callback.message, state)
    await callback.answer()


@router.callback_query(F.data.startswith("begin_test_"))
async def begin_test(callback: CallbackQuery, state: FSMContext):
    test_id = int(callback.data.split("_")[2])
    test = db.get_test(test_id)
    if not test:
        await callback.answer("Test topilmadi!", show_alert=True); return

    # Max urinish qayta tekshiruv
    max_att = test["max_attempts"] if test["max_attempts"] else 0
    if max_att > 0:
        attempt_count = db.get_user_test_attempt_count(callback.from_user.id, test_id)
        if attempt_count >= max_att:
            await callback.answer(
                f"⛔ Maksimal urinishlar soni ({max_att} ta) tugadi!", show_alert=True
            )
            return

    questions = db.get_questions(test_id, limit=test["question_count"])
    if not questions:
        await callback.answer("Bu testda savollar yo'q!", show_alert=True); return

    q_list = [dict(q) for q in questions]
    random.shuffle(q_list)

    await state.set_state(TestTakingState.taking_test)
    await state.update_data(
        test_id=test_id, questions=q_list, current=0, answers={},
        start_time=datetime.now().isoformat()
    )
    await callback.message.edit_text(
        f"✅ Test boshlandi!\n\n<b>{test['title']}</b>\n"
        f"⏱ Vaqt: {test['time_limit']} daqiqa\n"
        f"Savollar soni: {len(q_list)}\n\nBirinchi savol...",
        parse_mode=ParseMode.HTML
    )
    await callback.answer()
    await send_question(callback.message, state, callback.bot)


async def send_question(message: Message, state: FSMContext, bot: Bot = None):
    data = await state.get_data()
    questions = data.get("questions")
    current = data.get("current", 0)

    if not questions:
        await state.clear()
        try:
            await message.answer(
                "⚠️ Test sessiyasi tugagan. Iltimos, testni qaytadan boshlang.",
                reply_markup=main_menu
            )
        except Exception:
            pass
        return

    if current >= len(questions):
        await finish_test(message, state, bot)
        return

    q = questions[current]
    total = len(questions)
    progress = int((current / total) * 10)
    bar = "🟦" * progress + "⬜" * (10 - progress)

    text = (
        f"{bar}\n"
        f"<b>Savol {current + 1}/{total}</b>\n\n"
        f"❓ {q['question_text']}\n"
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=f"A) {q['option_a']}", callback_data="ans_A")],
        [InlineKeyboardButton(text=f"B) {q['option_b']}", callback_data="ans_B")],
        [InlineKeyboardButton(text=f"C) {q['option_c']}", callback_data="ans_C")],
        [InlineKeyboardButton(text=f"D) {q['option_d']}", callback_data="ans_D")],
        [InlineKeyboardButton(text="🛑 Testni to'xtatish", callback_data="stop_test")]
    ])
    try:
        if bot:
            await bot.send_message(message.chat.id, text, reply_markup=kb, parse_mode=ParseMode.HTML)
        else:
            await message.answer(text, reply_markup=kb, parse_mode=ParseMode.HTML)
    except Exception as e:
        logging.error(f"Savol yuborishda xato: {e}")


@router.callback_query(F.data.startswith("ans_"), TestTakingState.taking_test)
async def process_answer(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    questions = data["questions"]
    current = data["current"]
    answers = data["answers"]

    chosen = callback.data.split("_")[1]
    q = questions[current]
    correct = q["correct_answer"]
    answers[str(current)] = {"chosen": chosen, "correct": correct, "is_correct": chosen == correct}
    await state.update_data(answers=answers, current=current + 1)

    if chosen == correct:
        feedback = "✅ To'g'ri! +1"
    else:
        opt = {"A": q["option_a"], "B": q["option_b"], "C": q["option_c"], "D": q["option_d"]}
        feedback = f"❌ Noto'g'ri. To'g'ri: <b>{correct}) {opt[correct]}</b>"

    try:
        await callback.message.edit_text(callback.message.text + f"\n\n{feedback}", parse_mode=ParseMode.HTML)
    except Exception:
        pass
    await callback.answer()
    await asyncio.sleep(0.8)
    await send_question(callback.message, state, callback.bot)


@router.callback_query(F.data == "stop_test")
async def stop_test(callback: CallbackQuery, state: FSMContext):
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="Ha, to'xtataman", callback_data="confirm_stop"),
        InlineKeyboardButton(text="Yo'q, davom etaman", callback_data="continue_test")
    ]])
    await callback.message.edit_text(
        "⚠️ Testni to'xtatmoqchimisiz?\n\nTo'xtatilsa natijalar saqlanmaydi!",
        reply_markup=kb
    )
    await callback.answer()


@router.callback_query(F.data == "confirm_stop")
async def confirm_stop(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text("🛑 Test to'xtatildi.")
    await callback.message.answer("Asosiy menyu:", reply_markup=main_menu)
    await callback.answer()


@router.callback_query(F.data == "continue_test")
async def continue_test_cb(callback: CallbackQuery, state: FSMContext):
    await callback.answer("Davom eting! 💪")
    await send_question(callback.message, state, callback.bot)


async def finish_test(message: Message, state: FSMContext, bot: Bot = None):
    data = await state.get_data()
    questions = data.get("questions")
    answers = data.get("answers", {})
    test_id = data.get("test_id")

    if not questions or not test_id:
        await state.clear()
        try:
            await message.answer("⚠️ Test sessiyasi tugagan. Qaytadan boshlang.", reply_markup=main_menu)
        except Exception:
            pass
        return

    test = db.get_test(test_id)
    total = len(questions)
    score = sum(1 for a in answers.values() if a["is_correct"])
    percent = int(score / total * 100)
    chat_id = message.chat.id

    user_row = db.get_user(chat_id)
    db_name = user_row["full_name"] if user_row and user_row["full_name"] else (
        getattr(message.chat, "full_name", None) or str(chat_id)
    )

    passed, cert_code = db.save_result(chat_id, db_name, test_id, score, total)

    wrong_list = []
    for i, q in enumerate(questions):
        a = answers.get(str(i))
        if a and not a["is_correct"]:
            opt = {"A": q["option_a"], "B": q["option_b"], "C": q["option_c"], "D": q["option_d"]}
            wrong_list.append(
                f"• {q['question_text'][:50]}...\n"
                f"  Siz: {a['chosen']}) | To'g'ri: {a['correct']}) {opt[a['correct']]}"
            )

    result_text = (
        f"🏁 <b>Test yakunlandi!</b>\n\n"
        f"📋 Test: <b>{test['title']}</b>\n\n"
        f"📊 Natija: <b>{score}/{total}</b> ({percent}%)\n"
    )
    if passed:
        result_text += "\n🏆 <b>TABRIKLAYMIZ! Sertifikat oldingiz!</b>\n"
    else:
        need = int(CERTIFICATE_THRESHOLD * total) - score
        result_text += f"\n❌ Sertifikat uchun yana <b>{need} ta</b> to'g'ri javob kerak edi.\n"

    if wrong_list:
        result_text += f"\n📌 Noto'g'ri javoblar ({len(wrong_list)} ta):\n"
        result_text += "\n".join(wrong_list[:5])
        if len(wrong_list) > 5:
            result_text += f"\n... va yana {len(wrong_list) - 5} ta"

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔄 Qayta urinish", callback_data=f"start_test_{test_id}")],
        [InlineKeyboardButton(text="📝 Boshqa testlar", callback_data="back_to_tests")],
        [InlineKeyboardButton(text="🏠 Asosiy menyu", callback_data="back_to_main")]
    ])
    await state.clear()

    try:
        if bot:
            await bot.send_message(chat_id, result_text, reply_markup=kb, parse_mode=ParseMode.HTML)
        else:
            await message.answer(result_text, reply_markup=kb, parse_mode=ParseMode.HTML)

        if passed:
            cert_bytes = await generate_certificate(db_name, test["title"], score, total, cert_code)
            caption = (
                f"🎓 <b>Sertifikatingiz!</b>\n"
                f"🔖 Raqam: <code>{cert_code}</code>\n\n"
                f"QR kodni skanerlash orqali sertifikatni tekshirish mumkin.\n"
                f"Saqlang va ulashing 🌟"
            )
            if cert_bytes:
                cert_file = BufferedInputFile(cert_bytes, filename="sertifikat.png")
                if bot:
                    await bot.send_photo(chat_id, cert_file, caption=caption, parse_mode=ParseMode.HTML)
                else:
                    await message.answer_photo(cert_file, caption=caption, parse_mode=ParseMode.HTML)
            else:
                warn = (
                    f"🏆 <b>Tabriklaymiz!</b>\n\n"
                    f"🔖 Sertifikat raqami: <code>{cert_code}</code>\n"
                    f"Admin sertifikat shablonini yuklagandan so'ng avtomatik tayyorlanadi."
                )
                if bot:
                    await bot.send_message(chat_id, warn, parse_mode=ParseMode.HTML)
                else:
                    await message.answer(warn, parse_mode=ParseMode.HTML)
    except Exception as e:
        logging.error(f"Natija yuborishda xato: {e}", exc_info=True)


@router.callback_query(F.data == "my_results")
async def my_results(callback: CallbackQuery):
    results = db.get_user_results(callback.from_user.id)
    if not results:
        await callback.answer("Sizda hali natijalar yo'q!", show_alert=True); return
    text = "📊 <b>Mening natijalarim:</b>\n\n"
    cert_buttons = []
    for r in results[:10]:
        pct = int(r["score"] / r["total"] * 100) if r["total"] else 0
        status = "🏆 Sertifikat" if r["passed"] else "❌ O'tmadi"
        text += f"📋 {r['title']}\n{status} — {r['score']}/{r['total']} ({pct}%)\n\n"
        if r["passed"] and r["cert_code"]:
            cert_buttons.append([
                InlineKeyboardButton(
                    text=f"⬇️ {r['title'][:30]} sertifikatini yuklab olish",
                    callback_data=f"dl_cert_{r['cert_code']}"
                )
            ])
    inline_rows = cert_buttons + [
        [InlineKeyboardButton(text="⬅Orqaga", callback_data="back_to_tests")]
    ]
    kb = InlineKeyboardMarkup(inline_keyboard=inline_rows)
    await callback.message.edit_text(text, reply_markup=kb, parse_mode=ParseMode.HTML)
    await callback.answer()


@router.callback_query(F.data.startswith("dl_cert_"))
async def download_cert_callback(callback: CallbackQuery):
    cert_code = callback.data[len("dl_cert_"):]
    row = db.get_cert_by_code(cert_code)
    if not row:
        await callback.answer("❌ Sertifikat topilmadi!", show_alert=True)
        return
    if row["user_id"] != callback.from_user.id and not is_admin(callback.from_user.id):
        await callback.answer("❌ Bu sertifikat sizga tegishli emas!", show_alert=True)
        return
    await callback.answer("⏳ Sertifikat tayyorlanmoqda...")
    user = db.get_user(row["user_id"])
    full_name = user["full_name"] if user and user["full_name"] else row.get("user_name", "")
    cert_bytes = await generate_certificate(
        full_name, row["title"], row["score"], row["total"], cert_code
    )
    if cert_bytes:
        await callback.message.answer_document(
            BufferedInputFile(cert_bytes, filename=f"sertifikat_{cert_code}.png"),
            caption=(
                f"🏆 <b>Sertifikatingiz!</b>\n\n"
                f"📋 Test: <b>{row['title']}</b>\n"
                f"🔑 Kod: <code>{cert_code}</code>"
            ),
            parse_mode=ParseMode.HTML
        )
    else:
        await callback.message.answer(
            "⚠️ Sertifikat fayli yaratishda xatolik yuz berdi. Admin sertifikat shablonini yuklaganidan so'ng qayta urinib ko'ring."
        )


# =============================================================================
# ADMIN PANEL (faqat shaxsiy chat va adminlar uchun)
# =============================================================================
@router.message(F.text == "/admin")
async def admin_panel(message: Message):
    # Guruhlarda /admin ishlamaydi
    if message.chat.type in ("group", "supergroup"):
        return
    if not is_admin(message.from_user.id):
        await message.answer("❌ Sizga ruxsat yo'q!"); return
    await _send_admin_panel(message)


async def _send_admin_panel(message: Message, edit: bool = False, user_id: int = None):
    tests = db.get_all_tests()
    total_tests = len(tests)
    active_tests = sum(1 for t in tests if t["is_active"])
    kpi = db.get_kpi_stats()
    front_count = db.get_front_member_count()
    guv_count = db.get_card_count()
    appr_count = db.get_approved_count()
    text = (
        f"⚙️ <b>ADMIN PANEL</b>\n\n"
        f"👥 Foydalanuvchilar: <b>{kpi['total_users']}</b>  |  Bugun: <b>+{kpi['today_users']}</b>\n"
        f"📋 Testlar: <b>{total_tests}</b>  |  Faol: <b>{active_tests}</b>\n"
        f"🎯 Urinishlar: <b>{kpi['total_attempts']}</b>  |  Bugun: <b>{kpi['today_attempts']}</b>\n"
        f"🏆 Sertifikatlar: <b>{kpi['total_certs']}</b>  |  O'rtacha: <b>{kpi['avg_score']}%</b>\n"
        f"🏢 Front ofis a'zolari: <b>{front_count}</b>\n"
        f"🪪 Guvohnomalar: <b>{guv_count}</b>  |  Ro'yxatda: <b>{appr_count}</b>\n"
    )
    uid = user_id if user_id is not None else message.from_user.id
    rows = [
        [
            InlineKeyboardButton(text="Testlar", callback_data="admin_tests_list"),
            InlineKeyboardButton(text="Test yaratish", callback_data="admin_create_test"),
        ],
        [
            InlineKeyboardButton(text="Natijalar", callback_data="admin_results"),
            InlineKeyboardButton(text="KPI", callback_data="admin_kpi"),
        ],
        [
            InlineKeyboardButton(text="📊 Excel (barcha)", callback_data="admin_export_results"),
            InlineKeyboardButton(text="🏆 Excel (sertifikat)", callback_data="admin_export_certs"),
        ],
        [
            InlineKeyboardButton(text="📋 Excel (foydalanuvchilar)", callback_data="admin_export_users"),
            InlineKeyboardButton(text="Sertifikat shabloni", callback_data="admin_certificate_template"),
        ],
        [
            InlineKeyboardButton(text="🏢 Front ofis statistika", callback_data="admin_front_stats"),
        ],
        [
            InlineKeyboardButton(text="📥 Excel (front ofis a'zolari)", callback_data="admin_export_front"),
        ],
        [
            InlineKeyboardButton(text="🪪 A'zolik guvohnomasi", callback_data="admin_guv"),
        ],
        [
            InlineKeyboardButton(text="Broadcast", callback_data="admin_broadcast"),
            InlineKeyboardButton(text="Top foydalanuvchilar", callback_data="admin_top_users"),
        ],
    ]
    if is_superadmin(uid):
        rows.append([InlineKeyboardButton(text="Adminlar boshqaruvi", callback_data="admin_manage_admins")])
    kb = InlineKeyboardMarkup(inline_keyboard=rows)
    if edit:
        await safe_edit(message, text, reply_markup=kb)
    else:
        await message.answer(text, reply_markup=kb, parse_mode=ParseMode.HTML)


# ── FRONT OFIS: STATISTIKA VA EXCEL ──────────────────────────────────────────
@router.callback_query(F.data == "admin_front_stats")
async def admin_front_stats(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True); return
    stats = db.get_front_member_stats()

    text = (
        f"🏢 <b>FRONT OFIS A'ZOLARI STATISTIKASI</b>\n"
        f"{'─' * 30}\n"
        f"👥 Jami a'zolar: <b>{stats['total']}</b>\n"
        f"📅 Bugun qo'shilganlar: <b>+{stats['today']}</b>\n\n"
    )

    if stats["by_region"]:
        text += "🗺 <b>Hududlar bo'yicha:</b>\n"
        for row in stats["by_region"]:
            region = row["region"] or "Noma'lum"
            text += f"  • {region}: <b>{row['cnt']}</b>\n"
        text += "\n"

    if stats["by_category"]:
        text += "👤 <b>Yosh toifalari bo'yicha:</b>\n"
        for row in stats["by_category"]:
            cat = row["age_category"] or "Noma'lum"
            text += f"  • {cat}: <b>{row['cnt']}</b>\n"

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📥 Excel yuklab olish", callback_data="admin_export_front")],
        [InlineKeyboardButton(text="⬅️ Admin", callback_data="admin_back")]
    ])
    await safe_edit(callback.message, text, reply_markup=kb)
    await callback.answer()


@router.callback_query(F.data == "admin_export_front")
async def admin_export_front(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True); return
    await callback.answer("⏳ Excel tayyorlanmoqda...")

    members = db.get_all_front_members()
    if not members:
        await callback.message.answer("📭 Hali front ofis a'zolari yo'q.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Front ofis a'zolari"

    headers = [
        "#", "F.I.O", "Yosh", "Yosh toifasi", "Telefon",
        "Hudud", "Tuman/Shahar", "O'qish/Ish joyi",
        "Ro'yxatdan o'tgan sana", "Telegram ID"
    ]
    ws.append(headers)

    # Sarlavhalarni qalin qilish
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center")

    for i, m in enumerate(members, start=1):
        ws.append([
            i,
            m["full_name"] or "",
            m["age"] or "",
            m["age_category"] or "",
            m["phone"] or "",
            m["region"] or "",
            m["district"] or "",
            m["workplace"] or "",
            (m["registered_at"] or "")[:16],
            m["user_id"] or ""
        ])

    # Ustun kengliklarini sozlash
    widths = [5, 28, 6, 18, 16, 18, 20, 28, 18, 14]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"front_ofis_azolari_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    await callback.message.answer_document(
        BufferedInputFile(buf.read(), filename=filename),
        caption=f"🏢 <b>Front ofis a'zolari</b>\n\n👥 Jami: <b>{len(members)}</b> ta a'zo",
        parse_mode=ParseMode.HTML
    )


@router.callback_query(F.data == "admin_back")
async def admin_back_callback(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True); return
    await _send_admin_panel(callback.message, edit=True, user_id=callback.from_user.id)
    await callback.answer()


@router.callback_query(F.data == "admin_tests_list")
async def admin_tests_list(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True); return
    tests = db.get_all_tests()
    if not tests:
        await callback.message.edit_text(
            "📭 Hali testlar yo'q.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="➕ Test yaratish", callback_data="admin_create_test")],
                [InlineKeyboardButton(text="⬅️ Admin", callback_data="admin_back")]
            ])
        )
        await callback.answer(); return

    text = "📋 <b>Barcha testlar:</b>\n\n"
    buttons = []
    for t in tests:
        q_count = db.get_question_count(t["id"])
        status = "✅" if t["is_active"] else "⛔"
        max_att = t["max_attempts"] if t["max_attempts"] else 0
        att_str = f" | max: {max_att}" if max_att else ""
        text += f"{status} <b>{t['title']}</b> ({q_count} savol{att_str})\n"
        buttons.append([InlineKeyboardButton(
            text=f"{status} {t['title'][:30]}",
            callback_data=f"admin_test_detail_{t['id']}"
        )])
    buttons.append([InlineKeyboardButton(text="⬅️ Admin", callback_data="admin_back")])
    await callback.message.edit_text(
        text, reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons), parse_mode=ParseMode.HTML
    )
    await callback.answer()


async def _render_test_detail(callback: CallbackQuery, test_id: int):
    """Test detail sahifasini ko'rsatuvchi yordamchi funksiya."""
    test = db.get_test(test_id)
    if not test:
        await callback.answer("Test topilmadi!", show_alert=True); return
    q_count = db.get_question_count(test_id)
    status_text = "✅ Faol" if test["is_active"] else "⛔ Nofaol"
    toggle_text = "⛔ O'chirish" if test["is_active"] else "✅ Faollashtirish"
    max_att = test["max_attempts"] if test["max_attempts"] else 0
    max_att_str = f"{max_att} ta" if max_att else "Cheksiz"
    text = (
        f"<b>{test['title']}</b>\n\n"
        f"Tavsif: {test['description'] or 'Yoq'}\n"
        f"Savollar: {q_count} ta\n"
        f"Ko'rsatiladi: {test['question_count']} ta\n"
        f"Vaqt: {test['time_limit']} daqiqa\n"
        f"Max urinish: {max_att_str}\n"
        f"Holat: {status_text}\n"
        f"Yaratilgan: {test['created_at'][:10]}"
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text=toggle_text, callback_data=f"admin_toggle_{test_id}"),
            InlineKeyboardButton(text="Tahrirlash", callback_data=f"admin_edit_{test_id}")
        ],
        [
            InlineKeyboardButton(text="Savol (qo'lda)", callback_data=f"admin_add_q_{test_id}"),
            InlineKeyboardButton(text="Excel", callback_data=f"admin_excel_{test_id}")
        ],
        [
            InlineKeyboardButton(text="Natijalar", callback_data=f"admin_test_results_{test_id}"),
            InlineKeyboardButton(text="O'chirish", callback_data=f"admin_delete_{test_id}")
        ],
        [InlineKeyboardButton(text="⬅Orqaga", callback_data="admin_tests_list")]
    ])
    await callback.message.edit_text(text, reply_markup=kb, parse_mode=ParseMode.HTML)


@router.callback_query(F.data.startswith("admin_test_detail_"))
async def admin_test_detail(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True); return
    test_id = int(callback.data.split("_")[3])
    await _render_test_detail(callback, test_id)
    await callback.answer()


@router.callback_query(F.data.startswith("admin_toggle_"))
async def admin_toggle(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True); return
    test_id = int(callback.data.split("_")[2])
    new_status = db.toggle_test(test_id)
    msg = "✅ Faollashtirildi!" if new_status else "⛔ O'chirildi!"
    await callback.answer(msg, show_alert=True)
    await _render_test_detail(callback, test_id)


@router.callback_query(F.data.startswith("admin_delete_"))
async def admin_delete_test(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True); return
    test_id = int(callback.data.split("_")[2])
    test = db.get_test(test_id)
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="✅ Tasdiqlash", callback_data=f"admin_confirm_delete_{test_id}"),
            InlineKeyboardButton(text="❌ Bekor", callback_data=f"admin_test_detail_{test_id}")
        ]
    ])
    await callback.message.edit_text(
        f"⚠️ <b>'{test['title']}'</b> o'chirilsinmi?\nBarcha savollar ham o'chadi!",
        reply_markup=kb, parse_mode=ParseMode.HTML
    )
    await callback.answer()


@router.callback_query(F.data.startswith("admin_confirm_delete_"))
async def admin_confirm_delete(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True); return
    test_id = int(callback.data.split("_")[3])
    db.delete_test(test_id)
    await callback.answer("✅ O'chirildi!", show_alert=True)
    await admin_tests_list(callback)


@router.callback_query(F.data == "admin_create_test")
async def admin_create_test_start(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True); return
    await state.set_state(AdminTestState.creating_title)
    await callback.message.edit_text(
        "📝 <b>Yangi test yaratish</b>\n\nTest nomini kiriting:",
        parse_mode=ParseMode.HTML
    )
    await callback.answer()


@router.message(AdminTestState.creating_title, F.chat.type == "private")
async def admin_creating_title(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    await state.update_data(title=message.text)
    await state.set_state(AdminTestState.creating_description)
    await message.answer("Tavsif kiriting (yoki 'yoq' deb yozing):", reply_markup=cancel_kb)


@router.message(AdminTestState.creating_description, F.chat.type == "private")
async def admin_creating_desc(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    if message.text == "❌ Bekor qilish":
        await state.clear(); await message.answer("Bekor.", reply_markup=main_menu); return
    desc = "" if message.text.lower() in ["yoq", "yo'q"] else message.text
    await state.update_data(description=desc)
    await state.set_state(AdminTestState.creating_time)
    await message.answer("Vaqt limitini daqiqada kiriting (masalan: 30):", reply_markup=cancel_kb)


@router.message(RegisterState.waiting_for_birth_year, F.chat.type == "private")
async def process_birth_year(message: Message, state: FSMContext):
    try:
        year = int(message.text.strip())
        current_year = datetime.now().year
        if year < 1950 or year > current_year - 5:
            await message.answer(f"❗ Iltimos, haqiqiy tug'ilgan yilni kiriting (1950-{current_year - 5}):")
            return
    except ValueError:
        await message.answer("❗ Iltimos, faqat raqam kiriting (masalan: 2005):")
        return

    await state.update_data(birth_year=year)
    await state.set_state(RegisterState.waiting_for_region)
    await message.answer("🗺 Viloyatingizni tanlang:", reply_markup=region_kb)


@router.message(AdminTestState.creating_count, F.chat.type == "private")
async def admin_creating_count(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    if message.text == "❌ Bekor qilish":
        await state.clear(); await message.answer("Bekor.", reply_markup=main_menu); return
    try:
        c = int(message.text)
    except Exception:
        await message.answer("Raqam kiriting!"); return
    await state.update_data(question_count=c)
    await state.set_state(AdminTestState.creating_max_attempts)
    await message.answer(
        "🔄 <b>Maksimal urinishlar soni</b>\n\n"
        "Har bir foydalanuvchi bu testga necha marta urinishi mumkin?\n\n"
        "<i>0 yoki 'cheksiz' deb yozing — cheksiz urinishga ruxsat berish uchun</i>",
        parse_mode=ParseMode.HTML
    )


@router.message(AdminTestState.creating_max_attempts, F.chat.type == "private")
async def admin_creating_max_attempts(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    if message.text == "❌ Bekor qilish":
        await state.clear(); await message.answer("Bekor.", reply_markup=main_menu); return
    try:
        if message.text.lower() in ["cheksiz", "0", "∞"]:
            max_att = 0
        else:
            max_att = int(message.text)
            if max_att < 0:
                max_att = 0
    except Exception:
        await message.answer("Raqam kiriting (masalan: 3) yoki 'cheksiz' deb yozing!"); return

    data = await state.get_data()
    test_id = db.create_test(
        data["title"], data.get("description", ""),
        data["time_limit"], data["question_count"], max_att
    )
    await state.clear()
    max_str = f"{max_att} ta" if max_att else "Cheksiz"
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="➕ Savol qo'shish (qo'lda)", callback_data=f"admin_add_q_{test_id}")],
        [InlineKeyboardButton(text="📥 Excel orqali", callback_data=f"admin_excel_{test_id}")],
        [InlineKeyboardButton(text="📋 Ro'yxatga", callback_data="admin_tests_list")]
    ])
    await message.answer(
        f"✅ <b>Test yaratildi!</b>\n\n<b>{data['title']}</b>\n"
        f"🔄 Maksimal urinish: {max_str}\n\nEndi savol qo'shing:",
        reply_markup=kb, parse_mode=ParseMode.HTML
    )


@router.message(RegisterState.waiting_for_passport, F.chat.type == "private")
async def process_passport(message: Message, state: FSMContext):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Ro'yxatdan o'tish bekor qilindi. /start ni bosing.", reply_markup=main_menu)
        return

    passport = message.text.strip().upper() if message.text else ""
    if len(passport) < 5:
        await message.answer("❗ Pasport raqami juda qisqa. Iltimos, to'g'ri kiriting:", reply_markup=cancel_kb)
        return

    await state.update_data(passport=passport)
    await state.set_state(RegisterState.waiting_for_password)
    await message.answer(
        "🔑 <b>Parol o'rnatish</b>\n\n"
        "Test yechish uchun parol kiriting (kamida 4 ta belgi):\n"
        "<i>Bu parolni test yechishda kiritasiz.</i>",
        reply_markup=cancel_kb, parse_mode=ParseMode.HTML
    )


@router.message(AdminTestState.adding_q_a, F.chat.type == "private")
async def aq_a(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id): return
    if message.text == "❌ Bekor qilish":
        await state.clear(); await message.answer("Bekor.", reply_markup=main_menu); return
    await state.update_data(q_a=message.text)
    await state.set_state(AdminTestState.adding_q_b)
    await message.answer("B) variantini kiriting:", reply_markup=cancel_kb)


@router.message(TestTakingState.taking_test, F.chat.type == "private")
async def process_test_answer(message: Message, state: FSMContext, bot: Bot):
    """Test javoblarini qayta ishlash"""
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Test bekor qilindi.", reply_markup=main_menu)
        return
    data = await state.get_data()
    test_id = data.get("test_id")
    test = db.get_test(test_id)
    if not test:
        await message.answer("Test topilmadi!", reply_markup=main_menu)
        return
    q_count = db.get_question_count(test_id)
    if q_count <= 0:
        await message.answer("Testda savollar yo'q!", reply_markup=main_menu)
        return
    answers = data.get("answers", [])
    answers.append(message.text)
    await state.update_data(answers=answers)
    if len(answers) < q_count:
        await message.answer("Keyingi savol:", reply_markup=cancel_kb)
    else:
        await state.set_state(TestTakingState.confirming_stop)
        await message.answer(
            "✅ <b>Test tugadi!</b>\n\n"
            "Natijalaringiz:\n"
            "To'g'ri: {}\n"
            "Noto'g'ri: {}\n"
            "Umumiy: {}/{}\n\n"
            "Testni to'xtatib, natijalaringizni ko'rishni xohlaysizmi?".format(
                len([a for a in answers if a == "A"]),
                len([a for a in answers if a != "A"]),
                len(answers), q_count
            ),
            reply_markup=confirm_stop_kb, parse_mode=ParseMode.HTML
        )


@router.message(RegisterState.waiting_for_fullname, F.chat.type == "private")
async def process_fullname(message: Message, state: FSMContext):
    if not message.text or len(message.text.strip()) < 3:
        await message.answer("❗ Ism familya juda qisqa. Kamida 3 ta harf kiriting:")
        return
    await state.update_data(full_name=message.text.strip())
    await state.set_state(RegisterState.waiting_for_birth_year)
    await message.answer("📅 Tug'ilgan yilingizni kiriting (masalan: 2005):", reply_markup=cancel_kb)


@router.message(TestTakingState.confirming_stop, F.chat.type == "private")
async def confirm_stop_answer(message: Message, state: FSMContext, bot: Bot):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Davom etamiz!", reply_markup=main_menu)
        return
    data = await state.get_data()
    passport = data.get("passport")
    full_name = message.text
    user_id = message.from_user.id
    db.save_reset_request(user_id, passport, full_name)
    await message.answer(
        "⏳ <b>So'rov ketdi.</b>\n\nAdmin tekshirib javob beradi, kutib turin.",
        reply_markup=main_menu, parse_mode=ParseMode.HTML
    )
    await state.clear()
    admin_kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ Tiklash", callback_data=f"restore_{user_id}"),
        InlineKeyboardButton(text="❌ Bekor", callback_data=f"reject_{user_id}")
    ]])
    await notify_admins(
        bot,
        f"📢 <b>YANGI PAROL SO'ROVI!</b>\n\n"
        f"👤 <b>Kim:</b> {full_name}\n"
        f"📄 <b>Pasport:</b> {passport}\n"
        f"🆔 <b>ID:</b> {user_id}",
        reply_markup=admin_kb
    )


@router.callback_query(F.data.startswith("reject_"))
async def admin_reject(callback: CallbackQuery, bot: Bot):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer("Siz admin emassiz!", show_alert=True); return
    user_id = int(callback.data.split("_")[1])
    await callback.message.edit_text("❌ Bekor qilindi.")
    await bot.send_message(user_id, "❌ <b>Admin ruxsat bermadi.</b>", parse_mode=ParseMode.HTML)
    await callback.answer()


@router.callback_query(F.data.startswith("restore_"))
async def admin_restore_start(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer("Siz admin emassiz!", show_alert=True); return
    user_id = int(callback.data.split("_")[1])
    await state.update_data(target_user_id=user_id)
    await state.set_state(AdminState.waiting_for_new_password)
    await callback.message.edit_text("✍️ <b>Yangi parolni yozin:</b>", parse_mode=ParseMode.HTML)
    await callback.answer()


@router.message(AdminState.waiting_for_new_password, F.chat.type == "private")
async def admin_send_password(message: Message, state: FSMContext, bot: Bot):
    if message.from_user.id not in ADMIN_IDS: return
    data = await state.get_data()
    target_user_id = data.get("target_user_id")
    try:
        await bot.send_message(
            target_user_id,
            f"✅ <b>Parol tiklandi!</b>\n\n🔑 <b>Yangi paroliz:</b> <code>{message.text}</code>",
            parse_mode=ParseMode.HTML
        )
        await message.answer("✅ Userga ketti.")
    except Exception as e:
        await message.answer(f"Xato: {e}")
    await state.clear()


@router.message(F.text == "❓ Loyiha haqida savol berish", F.chat.type == "private")
async def ask_question_start(message: Message, state: FSMContext):
    await state.set_state(QuestionState.waiting_for_question)
    await message.answer("Savolingizni yozing:", reply_markup=cancel_kb)


@router.message(QuestionState.waiting_for_question, F.chat.type == "private")
async def process_question(message: Message, state: FSMContext, bot: Bot):
    if message.text != "❌ Bekor qilish":
        user_id = message.from_user.id
        user_name = message.from_user.full_name
        username = message.from_user.username or "yoq"
        db.clear_question_answered(user_id)
        admin_kb = InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="💬 Javob berish", callback_data=f"reply_user_{user_id}")
        ]])
        await notify_admins(
            bot,
            f"❓ <b>SAVOL KELDI!</b>\n\n"
            f"👤 {user_name}\n"
            f"🆔 <b>ID:</b> {user_id}\n"
            f"📱 @{username}\n\n"
            f"📝 {message.text}\n\n<i>Reply qiling yoki tugmani bosing!</i>",
            reply_markup=admin_kb
        )
        await message.answer("✅ Savolingiz adminga yuborildi!", reply_markup=main_menu)
    else:
        await message.answer("Bekor boldi", reply_markup=main_menu)
    await state.clear()


@router.message(F.text == "🚨 Xabar berish", F.chat.type == "private")
async def report_start(message: Message, bot: Bot):
    await notify_admins(
        bot,
        f"🚨 <b>STOP NARKO BOTGA O'TDI!</b>\n\n"
        f"👤 {message.from_user.full_name}\n"
        f"🆔 <code>{message.from_user.id}</code>\n"
        f"📱 @{message.from_user.username or 'yoq'}\n"
        f"⏱ {message.date.strftime('%Y-%m-%d %H:%M:%S')}"
    )
    await message.answer(
        f"🚨 <b>Xabar berish</b>\n\n<b>@{STOP_NARKO_BOT_USERNAME}</b> botiga o'ting:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📢 STOP NARKO BOT", url=f"https://t.me/{STOP_NARKO_BOT_USERNAME}")],
            [InlineKeyboardButton(text="🔙 Asosiy menyu", callback_data="back_to_main")]
        ]),
        parse_mode=ParseMode.HTML
    )


# =============================================================================
# FRONT OFIS A'ZOLIGIGA ARIZA
# =============================================================================
def _front_age_category_kb() -> InlineKeyboardMarkup:
    buttons = [
        [InlineKeyboardButton(text=cat, callback_data=f"fage_{i}")]
        for i, cat in enumerate(FRONT_AGE_CATEGORIES)
    ]
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def _front_region_kb() -> InlineKeyboardMarkup:
    buttons = []
    row = []
    for i, reg in enumerate(FRONT_REGIONS):
        row.append(InlineKeyboardButton(text=reg, callback_data=f"freg_{i}"))
        if len(row) == 2:
            buttons.append(row); row = []
    if row:
        buttons.append(row)
    return InlineKeyboardMarkup(inline_keyboard=buttons)


@router.message(F.text == "🏢 Front ofis a'zoligiga ariza", F.chat.type == "private")
async def front_apply_start(message: Message, state: FSMContext):
    await state.clear()
    if db.is_front_member(message.from_user.id):
        await message.answer(
            "✅ Siz allaqachon front ofis a'zoligiga ariza topshirgansiz!\n\n"
            "Adminlar arizangizni ko'rib chiqadi.",
            reply_markup=main_menu
        )
        return
    await state.set_state(FrontOfisState.waiting_fullname)
    await message.answer(
        "🏢 <b>FRONT OFIS A'ZOLIGIGA ARIZA</b>\n\n"
        "Front ofis a'zolari safiga qo'shilish uchun quyidagi ma'lumotlarni "
        "to'ldiring.\n\n"
        "✍️ <b>1/8.</b> To'liq <b>Ism Familiya (F.I.O)</b>ngizni kiriting:\n"
        "<i>(masalan: Alisher Karimov Akmal o'g'li)</i>",
        reply_markup=cancel_kb, parse_mode=ParseMode.HTML
    )


@router.message(FrontOfisState.waiting_fullname, F.chat.type == "private")
async def front_get_fullname(message: Message, state: FSMContext):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Ariza bekor qilindi.", reply_markup=main_menu)
        return
    if not message.text or len(message.text.strip()) < 3:
        await message.answer("❗ Iltimos, to'liq Ism Familiyangizni kiriting (kamida 3 ta harf):")
        return
    await state.update_data(full_name=message.text.strip())
    await state.set_state(FrontOfisState.waiting_age)
    await message.answer(
        "🎂 <b>2/8.</b> <b>Yoshingizni</b> kiriting (faqat raqam):\n"
        "<i>(masalan: 25)</i>",
        reply_markup=cancel_kb, parse_mode=ParseMode.HTML
    )


@router.message(FrontOfisState.waiting_age, F.chat.type == "private")
async def front_get_age(message: Message, state: FSMContext):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Ariza bekor qilindi.", reply_markup=main_menu)
        return
    try:
        age = int(message.text.strip())
        if age < 10 or age > 100:
            await message.answer("❗ Iltimos, haqiqiy yoshni kiriting (10-100):")
            return
    except (ValueError, AttributeError):
        await message.answer("❗ Iltimos, faqat raqam kiriting (masalan: 25):")
        return
    await state.update_data(age=age)
    await state.set_state(FrontOfisState.waiting_age_category)
    await message.answer(
        "👥 <b>3/8.</b> <b>Yosh toifangizni</b> tanlang:",
        reply_markup=_front_age_category_kb(), parse_mode=ParseMode.HTML
    )


@router.callback_query(F.data.startswith("fage_"), FrontOfisState.waiting_age_category)
async def front_get_age_category(callback: CallbackQuery, state: FSMContext):
    idx = int(callback.data.split("_")[1])
    if idx < 0 or idx >= len(FRONT_AGE_CATEGORIES):
        await callback.answer("Xato!"); return
    await state.update_data(age_category=FRONT_AGE_CATEGORIES[idx])
    await state.set_state(FrontOfisState.waiting_phone)
    await callback.message.edit_text(
        f"✅ Yosh toifasi: <b>{FRONT_AGE_CATEGORIES[idx]}</b>",
        parse_mode=ParseMode.HTML
    )
    await callback.message.answer(
        "📱 <b>4/8.</b> <b>Telefon raqamingizni</b> yuboring yoki kiriting:\n"
        "<i>(masalan: +998901234567)</i>",
        reply_markup=front_phone_kb, parse_mode=ParseMode.HTML
    )
    await callback.answer()


@router.message(FrontOfisState.waiting_phone, F.chat.type == "private")
async def front_get_phone(message: Message, state: FSMContext):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Ariza bekor qilindi.", reply_markup=main_menu)
        return
    phone = None
    if message.contact:
        phone = message.contact.phone_number
    elif message.text:
        phone = message.text.strip()
    if not phone:
        await message.answer("❗ Iltimos, telefon raqamingizni yuboring yoki kiriting:")
        return
    digits = phone.replace("+", "").replace(" ", "").replace("-", "")
    if not digits.isdigit() or len(digits) < 9:
        await message.answer("❗ Noto'g'ri format. Misol: +998901234567")
        return
    if not phone.startswith("+"):
        phone = "+" + phone if phone[0].isdigit() else phone
    await state.update_data(phone=phone)
    await state.set_state(FrontOfisState.waiting_region)
    await message.answer(
        "🗺 <b>5/8.</b> <b>Hududingizni (viloyat)</b> tanlang:",
        reply_markup=_front_region_kb(), parse_mode=ParseMode.HTML
    )


@router.callback_query(F.data.startswith("freg_"), FrontOfisState.waiting_region)
async def front_get_region(callback: CallbackQuery, state: FSMContext):
    idx = int(callback.data.split("_")[1])
    if idx < 0 or idx >= len(FRONT_REGIONS):
        await callback.answer("Xato!"); return
    await state.update_data(region=FRONT_REGIONS[idx])
    await state.set_state(FrontOfisState.waiting_district)
    await callback.message.edit_text(
        f"✅ Hudud: <b>{FRONT_REGIONS[idx]}</b>",
        parse_mode=ParseMode.HTML
    )
    await callback.message.answer(
        "🏘 <b>6/8.</b> <b>Tuman/Shahar</b> nomini kiriting:\n"
        "<i>(masalan: Chilonzor tumani)</i>",
        reply_markup=cancel_kb, parse_mode=ParseMode.HTML
    )
    await callback.answer()


@router.message(FrontOfisState.waiting_district, F.chat.type == "private")
async def front_get_district(message: Message, state: FSMContext):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Ariza bekor qilindi.", reply_markup=main_menu)
        return
    if not message.text or len(message.text.strip()) < 2:
        await message.answer("❗ Iltimos, Tuman/Shahar nomini kiriting:")
        return
    await state.update_data(district=message.text.strip())
    await state.set_state(FrontOfisState.waiting_workplace)
    await message.answer(
        "🏫 <b>7/8.</b> <b>O'qish/Ish joyingizni</b> kiriting:\n"
        "<i>(masalan: Toshkent Davlat Universiteti yoki \"Baxatech\" MChJ)</i>",
        reply_markup=cancel_kb, parse_mode=ParseMode.HTML
    )


@router.message(FrontOfisState.waiting_workplace, F.chat.type == "private")
async def front_get_workplace(message: Message, state: FSMContext):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Ariza bekor qilindi.", reply_markup=main_menu)
        return
    if not message.text or len(message.text.strip()) < 2:
        await message.answer("❗ Iltimos, O'qish/Ish joyingizni kiriting:")
        return
    await state.update_data(workplace=message.text.strip())
    await state.set_state(FrontOfisState.waiting_photo)
    await message.answer(
        "🖼 <b>8/8.</b> <b>3x4 rasmingizni</b> yuboring:\n"
        "<i>(Rasmni rasm/foto ko'rinishida yuboring)</i>",
        reply_markup=cancel_kb, parse_mode=ParseMode.HTML
    )


@router.message(FrontOfisState.waiting_photo, F.photo, F.chat.type == "private")
async def front_get_photo(message: Message, state: FSMContext, bot: Bot):
    photo_file_id = message.photo[-1].file_id
    data = await state.get_data()
    reg_date = datetime.now().strftime("%Y-%m-%d %H:%M")

    db.add_front_member(
        user_id=message.from_user.id,
        full_name=data.get("full_name"),
        age=data.get("age"),
        age_category=data.get("age_category"),
        phone=data.get("phone"),
        region=data.get("region"),
        district=data.get("district"),
        workplace=data.get("workplace"),
        photo_file_id=photo_file_id
    )
    await state.clear()

    await message.answer(
        "✅ <b>Arizangiz qabul qilindi!</b>\n\n"
        "🏢 Front ofis a'zoligiga ariza topshirganingiz uchun rahmat.\n"
        "Adminlar arizangizni ko'rib chiqadi va tez orada bog'lanadi.",
        reply_markup=main_menu, parse_mode=ParseMode.HTML
    )

    # Adminlarga xabar berish
    admin_caption = (
        f"🏢 <b>YANGI FRONT OFIS ARIZASI!</b>\n"
        f"{'─' * 30}\n"
        f"👤 <b>F.I.O:</b> {data.get('full_name')}\n"
        f"🎂 <b>Yosh:</b> {data.get('age')}\n"
        f"👥 <b>Yosh toifasi:</b> {data.get('age_category')}\n"
        f"📱 <b>Telefon:</b> {data.get('phone')}\n"
        f"🗺 <b>Hudud:</b> {data.get('region')}\n"
        f"🏘 <b>Tuman/Shahar:</b> {data.get('district')}\n"
        f"🏫 <b>O'qish/Ish joyi:</b> {data.get('workplace')}\n"
        f"📅 <b>Sana:</b> {reg_date}\n"
        f"🆔 <b>ID:</b> {message.from_user.id}"
    )
    for aid in get_all_admin_ids():
        try:
            await bot.send_photo(
                aid, photo_file_id, caption=admin_caption,
                parse_mode=ParseMode.HTML
            )
        except Exception:
            pass


@router.message(FrontOfisState.waiting_photo, F.chat.type == "private")
async def front_photo_invalid(message: Message, state: FSMContext):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Ariza bekor qilindi.", reply_markup=main_menu)
        return
    await message.answer("❗ Iltimos, 3x4 rasmingizni rasm (foto) ko'rinishida yuboring.")


# =============================================================================
# A'ZOLIK GUVOHNOMASI — FOYDALANUVCHI
# =============================================================================
guv_phone_kb = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="📱 Telefon raqamni yuborish", request_contact=True)],
        [KeyboardButton(text="❌ Bekor qilish")]
    ],
    resize_keyboard=True
)

guv_skip_kb = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="⏭ O'tkazib yuborish")],
        [KeyboardButton(text="❌ Bekor qilish")]
    ],
    resize_keyboard=True
)


def _guv_region_kb() -> InlineKeyboardMarkup:
    buttons, row = [], []
    for code, name in GUV_REGIONS:
        row.append(InlineKeyboardButton(text=name, callback_data=f"guvreg_{code}"))
        if len(row) == 2:
            buttons.append(row); row = []
    if row:
        buttons.append(row)
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def _guv_card_text(card) -> str:
    fio = " ".join(x for x in [card["last_name"], card["first_name"],
                               card["middle_name"]] if x)
    return (
        f"🪪 <b>A'ZOLIK GUVOHNOMASI</b>\n"
        f"{'─' * 28}\n"
        f"👤 <b>F.I.Sh:</b> {fio}\n"
        f"🆔 <b>ID raqami:</b> <code>{card['card_id']}</code>\n"
        f"🔖 <b>Seriya raqami:</b> <code>{card['serial']}</code>\n"
        f"🗺 <b>Hudud:</b> {card['region']}\n"
        f"💼 <b>Lavozim:</b> {card['position']}\n"
        f"📅 <b>Berilgan sana:</b> {card['issued_date']}\n"
        f"⏳ <b>Amal qilish muddati:</b> {card['expires_date']}"
    )


async def _guv_download_photo(bot: Bot, file_id: str) -> bytes | None:
    if not file_id:
        return None
    try:
        f = await bot.get_file(file_id)
        buf = io.BytesIO()
        await bot.download_file(f.file_path, buf)
        return buf.getvalue()
    except Exception as e:
        logging.warning(f"[GUV] Rasm yuklab olinmadi: {e}")
        return None


async def _guv_render(bot: Bot, card):
    """Bazadagi yozuvdan guvohnomaning ikki tomonini qayta chizadi."""
    photo_bytes = await _guv_download_photo(bot, card["photo_file_id"])
    data = {
        "card_id": card["card_id"],
        "serial": card["serial"],
        "last_name": card["last_name"],
        "first_name": card["first_name"],
        "middle_name": card["middle_name"] or "",
        "issued_date": card["issued_date"],
        "position": card["position"],
    }
    return await generate_membership_card(data, photo_bytes)


async def _guv_send(message: Message, bot: Bot, card, as_document: bool = False):
    front_b, back_b = await _guv_render(bot, card)
    if not front_b:
        await message.answer(
            "❌ Guvohnomani tayyorlashda xatolik yuz berdi. "
            "Iltimos, keyinroq urinib ko'ring.", reply_markup=main_menu
        )
        return
    base = f"guvohnoma_{card['card_id']}"
    if as_document:
        await message.answer_document(
            BufferedInputFile(front_b, filename=f"{base}_old.png"),
            caption="🖨 Old tomon (yuqori sifat)"
        )
        await message.answer_document(
            BufferedInputFile(back_b, filename=f"{base}_orqa.png"),
            caption="🖨 Orqa tomon (yuqori sifat)"
        )
        return

    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="🖨 Chop etish uchun (yuqori sifat)",
                             callback_data="guv_hq")
    ]])
    await message.answer_media_group(media=[
        InputMediaPhoto(media=BufferedInputFile(front_b, filename=f"{base}_old.png"),
                        caption=_guv_card_text(card), parse_mode=ParseMode.HTML),
        InputMediaPhoto(media=BufferedInputFile(back_b, filename=f"{base}_orqa.png")),
    ])
    await message.answer(
        "✅ Guvohnomangiz tayyor! Orqa tomondagi QR kod orqali "
        "istalgan vaqtda haqiqiyligini tekshirish mumkin.",
        reply_markup=kb
    )


@router.message(F.text == "🪪 A'zolik guvohnomasi", F.chat.type == "private")
async def guv_start(message: Message, state: FSMContext, bot: Bot):
    await state.clear()
    uid = message.from_user.id

    card = db.get_card_by_user(uid)
    if card:
        await message.answer("🪪 Sizda a'zolik guvohnomasi mavjud:",
                             reply_markup=main_menu)
        await _guv_send(message, bot, card)
        return

    rec = db.find_approved_by_tg(uid)
    if rec is None:
        await state.set_state(GuvState.waiting_phone)
        await state.update_data(uid=uid)
        await message.answer(
            "🪪 <b>A'ZOLIK GUVOHNOMASI</b>\n\n"
            "Guvohnoma faqat <b>tasdiqlangan a'zolar ro'yxatida</b> "
            "bo'lgan shaxslarga beriladi.\n\n"
            "📱 Sizni ro'yxatdan topishimiz uchun <b>telefon raqamingizni</b> "
            "yuboring:",
            reply_markup=guv_phone_kb, parse_mode=ParseMode.HTML
        )
        return

    await _guv_load_record(state, uid, rec)
    await message.answer(
        "🪪 <b>A'ZOLIK GUVOHNOMASI</b>\n\n"
        "✅ Siz tasdiqlangan a'zolar ro'yxatidasiz.",
        parse_mode=ParseMode.HTML
    )
    await _guv_next(message, state, bot)


async def _guv_load_record(state: FSMContext, uid: int, rec):
    code = guv_region_code(rec["region"]) if rec["region"] else None
    await state.update_data(
        uid=uid,
        appr_id=rec["id"],
        last_name=guv_norm_text(rec["last_name"]) or None,
        first_name=guv_norm_text(rec["first_name"]) or None,
        middle_name=guv_norm_text(rec["middle_name"]) if rec["middle_name"] else None,
        region_code=code,
        position=guv_norm_text(rec["position"]) if rec["position"] else None,
        phone=rec["phone"],
    )


async def _guv_next(message: Message, state: FSMContext, bot: Bot):
    """Yetishmayotgan ma'lumotni so'raydi yoki guvohnomani beradi."""
    d = await state.get_data()

    if d.get("last_name") is None:
        await state.set_state(GuvState.waiting_last_name)
        return await message.answer(
            "✍️ <b>Familiyangizni</b> kiriting:\n<i>(masalan: Bahodirov)</i>",
            reply_markup=cancel_kb, parse_mode=ParseMode.HTML)

    if d.get("first_name") is None:
        await state.set_state(GuvState.waiting_first_name)
        return await message.answer(
            "✍️ <b>Ismingizni</b> kiriting:\n<i>(masalan: Baxtiyorjon)</i>",
            reply_markup=cancel_kb, parse_mode=ParseMode.HTML)

    if d.get("middle_name") is None:
        await state.set_state(GuvState.waiting_middle_name)
        return await message.answer(
            "✍️ <b>Otangizning ismini</b> kiriting:\n"
            "<i>(masalan: Abdulaziz o'g'li)</i>",
            reply_markup=guv_skip_kb, parse_mode=ParseMode.HTML)

    if d.get("region_code") is None:
        await state.set_state(GuvState.waiting_region)
        return await message.answer(
            "🗺 <b>Hududingizni</b> tanlang:\n"
            "<i>(ID raqamingizning birinchi ikki raqami shu hududga tegishli bo'ladi)</i>",
            reply_markup=_guv_region_kb(), parse_mode=ParseMode.HTML)

    if d.get("photo_file_id") is None:
        uid = d.get("uid") or message.chat.id
        fm = db.get_front_member(uid)
        if fm and fm["photo_file_id"]:
            await state.update_data(photo_file_id=fm["photo_file_id"])
            return await _guv_issue(message, state, bot)
        await state.set_state(GuvState.waiting_photo)
        return await message.answer(
            "🖼 <b>3x4 rasmingizni</b> yuboring:\n"
            "<i>(Oq fonda, yuzingiz aniq ko'rinadigan rasm bo'lsin)</i>",
            reply_markup=cancel_kb, parse_mode=ParseMode.HTML)

    await _guv_issue(message, state, bot)


async def _guv_issue(message: Message, state: FSMContext, bot: Bot):
    d = await state.get_data()
    uid = d.get("uid") or message.chat.id
    wait = await message.answer("⏳ Guvohnoma tayyorlanmoqda...")

    code = d["region_code"]
    region = GUV_REGION_NAME.get(code, "")
    position = d.get("position") or f"{region} koordinatori"
    card_id = guv_make_card_id(code)

    db.save_membership_card(
        card_id=card_id, user_id=uid,
        last_name=guv_norm_text(d["last_name"]),
        first_name=guv_norm_text(d["first_name"]),
        middle_name=guv_norm_text(d.get("middle_name") or ""),
        region=region, region_code=code, position=position,
        phone=d.get("phone"), photo_file_id=d.get("photo_file_id"),
        serial=card_id[-4:],
        issued_date=datetime.now().strftime("%d.%m.%Y"),
        expires_date=GUV_EXPIRE_DATE,
    )
    await state.clear()

    try:
        await wait.delete()
    except Exception:
        pass

    card = db.get_card_by_id(card_id)
    await message.answer("✅ <b>Guvohnomangiz rasmiylashtirildi!</b>",
                         reply_markup=main_menu, parse_mode=ParseMode.HTML)
    await _guv_send(message, bot, card)

    for aid in get_all_admin_ids():
        try:
            await bot.send_message(
                aid,
                f"🪪 <b>YANGI GUVOHNOMA BERILDI</b>\n{'─' * 28}\n"
                f"{_guv_card_text(card)}\n\n🆔 <b>Telegram ID:</b> <code>{uid}</code>",
                parse_mode=ParseMode.HTML
            )
        except Exception:
            pass


@router.callback_query(F.data == "guv_hq")
async def guv_high_quality(callback: CallbackQuery, bot: Bot):
    card = db.get_card_by_user(callback.from_user.id)
    if not card:
        await callback.answer("Guvohnoma topilmadi.", show_alert=True); return
    await callback.answer("⏳ Tayyorlanmoqda...")
    await _guv_send(callback.message, bot, card, as_document=True)


# ── BOSQICHLAR ───────────────────────────────────────────────────────────────
@router.message(GuvState.waiting_phone, F.chat.type == "private")
async def guv_get_phone(message: Message, state: FSMContext, bot: Bot):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Bekor qilindi.", reply_markup=main_menu); return

    phone = message.contact.phone_number if message.contact else (message.text or "")
    key = guv_phone_key(phone)
    if not key:
        await message.answer("❗ Noto'g'ri format. Masalan: <code>+998901234567</code>",
                             parse_mode=ParseMode.HTML)
        return

    uid = message.from_user.id
    rec = db.find_approved_by_phone(key)
    if rec is None:
        await state.clear()
        await message.answer(
            "❌ <b>Siz tasdiqlangan a'zolar ro'yxatida topilmadingiz.</b>\n\n"
            "Guvohnoma faqat admin tomonidan yuklangan a'zolar tarkibidagi "
            "shaxslarga beriladi.\n\n"
            "Agar a'zo bo'lsangiz — hududingiz koordinatoriga murojaat qiling "
            "yoki <b>🏢 Front ofis a'zoligiga ariza</b> bo'limi orqali ariza topshiring.",
            reply_markup=main_menu, parse_mode=ParseMode.HTML
        )
        return

    db.bind_approved_tg(rec["id"], uid)
    await _guv_load_record(state, uid, rec)
    if not (await state.get_data()).get("phone"):
        await state.update_data(phone=phone)
    await message.answer("✅ Topildingiz! Siz tasdiqlangan a'zolar ro'yxatidasiz.",
                         reply_markup=cancel_kb)
    await _guv_next(message, state, bot)


@router.message(GuvState.waiting_last_name, F.text, F.chat.type == "private")
async def guv_get_last_name(message: Message, state: FSMContext, bot: Bot):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Bekor qilindi.", reply_markup=main_menu); return
    val = guv_norm_text(message.text)
    if len(val) < 2:
        await message.answer("❗ Familiya kamida 2 ta harfdan iborat bo'lsin:"); return
    await state.update_data(last_name=val)
    await _guv_next(message, state, bot)


@router.message(GuvState.waiting_first_name, F.text, F.chat.type == "private")
async def guv_get_first_name(message: Message, state: FSMContext, bot: Bot):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Bekor qilindi.", reply_markup=main_menu); return
    val = guv_norm_text(message.text)
    if len(val) < 2:
        await message.answer("❗ Ism kamida 2 ta harfdan iborat bo'lsin:"); return
    await state.update_data(first_name=val)
    await _guv_next(message, state, bot)


@router.message(GuvState.waiting_middle_name, F.text, F.chat.type == "private")
async def guv_get_middle_name(message: Message, state: FSMContext, bot: Bot):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Bekor qilindi.", reply_markup=main_menu); return
    val = "" if message.text == "⏭ O'tkazib yuborish" else guv_norm_text(message.text)
    await state.update_data(middle_name=val)
    await _guv_next(message, state, bot)


@router.callback_query(F.data.startswith("guvreg_"), GuvState.waiting_region)
async def guv_set_region(callback: CallbackQuery, state: FSMContext, bot: Bot):
    code = callback.data.split("_", 1)[1]
    if code not in GUV_REGION_NAME:
        await callback.answer("Xato!"); return
    await state.update_data(region_code=code)
    await safe_edit(callback.message,
                    f"✅ Hudud: <b>{GUV_REGION_NAME[code]}</b> (kod: <code>{code}</code>)")
    await callback.answer()
    await _guv_next(callback.message, state, bot)


@router.message(GuvState.waiting_photo, F.photo, F.chat.type == "private")
async def guv_get_photo(message: Message, state: FSMContext, bot: Bot):
    await state.update_data(photo_file_id=message.photo[-1].file_id)
    await _guv_next(message, state, bot)


@router.message(GuvState.waiting_photo, F.chat.type == "private")
async def guv_photo_invalid(message: Message, state: FSMContext):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Bekor qilindi.", reply_markup=main_menu); return
    await message.answer("❗ Iltimos, 3x4 rasmni <b>rasm (foto)</b> ko'rinishida yuboring.",
                         parse_mode=ParseMode.HTML)


# =============================================================================
# A'ZOLIK GUVOHNOMASI — ADMIN
# =============================================================================
_GUV_COL_MAP = [
    ("middle_name", ("otasining", "otasi", "sharif", "otchestvo", "middle")),
    ("last_name",   ("familiya", "familya", "surname", "last name", "familiyasi")),
    ("first_name",  ("ismi", "ism", "name", "imya")),
    ("tg_id",       ("telegram", "tg id", "tgid", "chat id", "user id", "id raqam")),
    ("phone",       ("telefon", "phone", "raqam", "nomer", "tel")),
    ("region",      ("hudud", "viloyat", "region", "shahar")),
    ("position",    ("lavozim", "position", "vazifa")),
]
_GUV_DEFAULT_ORDER = ["last_name", "first_name", "middle_name",
                      "tg_id", "phone", "region", "position"]


def _guv_detect_columns(header_row) -> dict:
    """Sarlavha qatoridan ustunlarni aniqlaydi; topilmasa tartib bo'yicha oladi."""
    mapping = {}
    for idx, cell in enumerate(header_row):
        h = str(cell or "").strip().lower()
        if not h:
            continue
        for field, keys in _GUV_COL_MAP:
            if field in mapping.values():
                continue
            if any(k in h for k in keys):
                mapping[idx] = field
                break
    if len(mapping) < 2:
        mapping = {i: f for i, f in enumerate(_GUV_DEFAULT_ORDER)}
    return mapping


def _guv_admin_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📥 Ro'yxatni yuklash (Excel)",
                              callback_data="admin_guv_upload")],
        [InlineKeyboardButton(text="📄 Excel namuna",
                              callback_data="admin_guv_sample")],
        [InlineKeyboardButton(text="📊 Ro'yxat (Excel)",
                              callback_data="admin_guv_export_list"),
         InlineKeyboardButton(text="🪪 Berilganlar (Excel)",
                              callback_data="admin_guv_export_cards")],
        [InlineKeyboardButton(text="🗑 Ro'yxatni tozalash",
                              callback_data="admin_guv_clear")],
        [InlineKeyboardButton(text="⬅️ Admin", callback_data="admin_back")],
    ])


@router.callback_query(F.data == "admin_guv")
async def admin_guv_menu(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True); return
    await state.clear()
    await safe_edit(
        callback.message,
        f"🪪 <b>A'ZOLIK GUVOHNOMASI</b>\n{'─' * 28}\n"
        f"📋 Tasdiqlangan a'zolar ro'yxati: <b>{db.get_approved_count()}</b>\n"
        f"🪪 Berilgan guvohnomalar: <b>{db.get_card_count()}</b>\n\n"
        f"Faqat ro'yxatga yuklangan shaxslar guvohnoma ola oladi.",
        reply_markup=_guv_admin_kb()
    )
    await callback.answer()


@router.callback_query(F.data == "admin_guv_sample")
async def admin_guv_sample(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True); return
    await callback.answer("⏳ Namuna tayyorlanmoqda...")

    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "A'zolar"
    headers = ["Familiyasi", "Ismi", "Otasining ismi", "Telegram ID",
               "Telefon", "Hudud", "Lavozim (ixtiyoriy)"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True); c.alignment = Alignment(horizontal="center")
    ws.append(["Bahodirov", "Baxtiyorjon", "Abdulaziz o'g'li",
               2110945697, "+998901234567", "Toshkent shahri", ""])
    ws.append(["Karimova", "Nilufar", "Akmal qizi",
               "", "+998935556677", "Farg'ona", "Farg'ona viloyati koordinatori"])
    for i, w in enumerate([20, 18, 22, 16, 18, 22, 30], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)

    await callback.message.answer_document(
        BufferedInputFile(buf.read(), filename="azolar_royxati_namuna.xlsx"),
        caption=(
            "📄 <b>A'zolar ro'yxati namunasi</b>\n\n"
            "• <b>Telegram ID</b> yoki <b>Telefon</b> — kamida bittasi bo'lishi shart\n"
            "• <b>Hudud</b> ID raqamning birinchi 2 raqamini belgilaydi\n"
            "• <b>Lavozim</b> bo'sh bo'lsa — «&lt;Hudud&gt; koordinatori» yoziladi\n\n"
            "Shu ustunlarni to'ldirib, «📥 Ro'yxatni yuklash» orqali yuboring."
        ),
        parse_mode=ParseMode.HTML
    )


@router.callback_query(F.data == "admin_guv_upload")
async def admin_guv_upload_start(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True); return
    await state.set_state(AdminGuvState.waiting_excel)
    await callback.message.answer(
        "📥 <b>A'zolar ro'yxatini yuklash</b>\n\n"
        "Excel faylni (<code>.xlsx</code>) shu yerga yuboring.\n"
        "Ustunlar: Familiyasi | Ismi | Otasining ismi | Telegram ID | "
        "Telefon | Hudud | Lavozim\n\n"
        "Mavjud a'zolar (Telegram ID yoki telefon bo'yicha) yangilanadi.",
        reply_markup=cancel_kb, parse_mode=ParseMode.HTML
    )
    await callback.answer()


@router.message(AdminGuvState.waiting_excel, F.document, F.chat.type == "private")
async def admin_guv_upload_file(message: Message, state: FSMContext, bot: Bot):
    if not is_admin(message.from_user.id):
        await state.clear(); return
    name = (message.document.file_name or "").lower()
    if not name.endswith((".xlsx", ".xlsm")):
        await message.answer("❗ Faqat <code>.xlsx</code> fayl qabul qilinadi.",
                             parse_mode=ParseMode.HTML)
        return

    await state.clear()
    wait = await message.answer("⏳ Fayl o'qilmoqda...")
    try:
        f = await bot.get_file(message.document.file_id)
        buf = io.BytesIO()
        await bot.download_file(f.file_path, buf)
        buf.seek(0)
        wb = openpyxl.load_workbook(buf, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        logging.error(f"[GUV] Excel xato: {e}", exc_info=True)
        await wait.edit_text("❌ Faylni o'qib bo'lmadi. Formatni tekshiring.")
        return

    if not rows:
        await wait.edit_text("❌ Fayl bo'sh."); return

    colmap = _guv_detect_columns(rows[0])
    header_is_text = any(
        isinstance(c, str) and any(k in str(c).lower() for _, ks in _GUV_COL_MAP for k in ks)
        for c in rows[0]
    )
    data_rows = rows[1:] if header_is_text else rows

    added = updated = skipped = 0
    bad_region = []
    for i, r in enumerate(data_rows, start=2):
        vals = {}
        for idx, field in colmap.items():
            vals[field] = r[idx] if idx < len(r) else None

        tg_id = None
        raw_tg = str(vals.get("tg_id") or "").strip()
        if raw_tg:
            digits = re.sub(r"\D", "", raw_tg)
            if digits:
                tg_id = int(digits)

        phone = str(vals.get("phone") or "").strip()
        if phone.endswith(".0"):
            phone = phone[:-2]
        key = guv_phone_key(phone)
        if phone and not phone.startswith("+") and key:
            phone = "+998" + key if len(re.sub(r"\D", "", phone)) <= 9 else "+" + re.sub(r"\D", "", phone)

        last = guv_norm_text(str(vals.get("last_name") or "")) or None
        first = guv_norm_text(str(vals.get("first_name") or "")) or None
        middle = guv_norm_text(str(vals.get("middle_name") or "")) or None
        region_raw = str(vals.get("region") or "").strip()
        position = guv_norm_text(str(vals.get("position") or "")) or None

        if not tg_id and not key:
            skipped += 1
            continue
        if not last and not first:
            skipped += 1
            continue

        region_name = None
        if region_raw:
            code = guv_region_code(region_raw)
            if code:
                region_name = GUV_REGION_NAME[code]
            else:
                bad_region.append(f"{i}-qator: «{region_raw}»")

        res = db.upsert_approved_member(
            tg_id=tg_id, phone_key=key or None, phone=phone or None,
            last_name=last, first_name=first, middle_name=middle,
            region=region_name, position=position,
            added_by=message.from_user.id
        )
        if res == "added":
            added += 1
        else:
            updated += 1

    text = (
        f"✅ <b>Ro'yxat yuklandi</b>\n{'─' * 28}\n"
        f"➕ Yangi qo'shildi: <b>{added}</b>\n"
        f"♻️ Yangilandi: <b>{updated}</b>\n"
        f"⏭ O'tkazib yuborildi: <b>{skipped}</b>\n"
        f"📋 Jami ro'yxatda: <b>{db.get_approved_count()}</b>"
    )
    if bad_region:
        text += ("\n\n⚠️ <b>Hudud tanilmadi</b> (bo'sh qoldirildi, "
                 "foydalanuvchidan so'raladi):\n• " + "\n• ".join(bad_region[:10]))
        if len(bad_region) > 10:
            text += f"\n• ... yana {len(bad_region) - 10} ta"

    await wait.delete()
    await message.answer(text, reply_markup=main_menu, parse_mode=ParseMode.HTML)


@router.message(AdminGuvState.waiting_excel, F.chat.type == "private")
async def admin_guv_upload_invalid(message: Message, state: FSMContext):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Bekor qilindi.", reply_markup=main_menu); return
    await message.answer("❗ Iltimos, <code>.xlsx</code> faylni yuboring.",
                         parse_mode=ParseMode.HTML)


@router.callback_query(F.data == "admin_guv_export_list")
async def admin_guv_export_list(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True); return
    rows = db.get_all_approved()
    if not rows:
        await callback.answer("📭 Ro'yxat bo'sh.", show_alert=True); return
    await callback.answer("⏳ Excel tayyorlanmoqda...")

    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Tasdiqlangan a'zolar"
    ws.append(["#", "Familiyasi", "Ismi", "Otasining ismi", "Telegram ID",
               "Telefon", "Hudud", "Lavozim", "Qo'shilgan sana"])
    for c in ws[1]:
        c.font = Font(bold=True); c.alignment = Alignment(horizontal="center")
    for i, m in enumerate(rows, start=1):
        ws.append([i, m["last_name"] or "", m["first_name"] or "",
                   m["middle_name"] or "", m["tg_id"] or "", m["phone"] or "",
                   m["region"] or "", m["position"] or "",
                   (m["added_at"] or "")[:16]])
    for i, w in enumerate([5, 20, 18, 22, 16, 18, 22, 30, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    await callback.message.answer_document(
        BufferedInputFile(buf.read(),
                          filename=f"guvohnoma_royxat_{datetime.now():%Y%m%d_%H%M}.xlsx"),
        caption=f"📋 Tasdiqlangan a'zolar: <b>{len(rows)}</b>",
        parse_mode=ParseMode.HTML
    )


@router.callback_query(F.data == "admin_guv_export_cards")
async def admin_guv_export_cards(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True); return
    cards = db.get_all_cards()
    if not cards:
        await callback.answer("📭 Hali guvohnoma berilmagan.", show_alert=True); return
    await callback.answer("⏳ Excel tayyorlanmoqda...")

    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Guvohnomalar"
    ws.append(["#", "ID raqami", "Seriya", "Familiyasi", "Ismi", "Otasining ismi",
               "Hudud", "Lavozim", "Telefon", "Berilgan sana",
               "Amal muddati", "Telegram ID"])
    for c in ws[1]:
        c.font = Font(bold=True); c.alignment = Alignment(horizontal="center")
    for i, m in enumerate(cards, start=1):
        ws.append([i, m["card_id"], m["serial"], m["last_name"] or "",
                   m["first_name"] or "", m["middle_name"] or "", m["region"] or "",
                   m["position"] or "", m["phone"] or "", m["issued_date"] or "",
                   m["expires_date"] or "", m["user_id"] or ""])
    for i, w in enumerate([5, 12, 10, 20, 18, 22, 22, 30, 18, 14, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    await callback.message.answer_document(
        BufferedInputFile(buf.read(),
                          filename=f"guvohnomalar_{datetime.now():%Y%m%d_%H%M}.xlsx"),
        caption=f"🪪 Berilgan guvohnomalar: <b>{len(cards)}</b>",
        parse_mode=ParseMode.HTML
    )


@router.callback_query(F.data == "admin_guv_clear")
async def admin_guv_clear(callback: CallbackQuery):
    if not is_superadmin(callback.from_user.id):
        await callback.answer("❌ Faqat superadmin!", show_alert=True); return
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🗑 Ha, tozalansin",
                              callback_data="admin_guv_clear_yes"),
         InlineKeyboardButton(text="⬅️ Yo'q", callback_data="admin_guv")]
    ])
    await safe_edit(callback.message,
                    "⚠️ <b>Tasdiqlangan a'zolar ro'yxati butunlay o'chiriladi.</b>\n\n"
                    "Berilgan guvohnomalar saqlanib qoladi. Davom etamizmi?",
                    reply_markup=kb)
    await callback.answer()


@router.callback_query(F.data == "admin_guv_clear_yes")
async def admin_guv_clear_yes(callback: CallbackQuery):
    if not is_superadmin(callback.from_user.id):
        await callback.answer("❌ Faqat superadmin!", show_alert=True); return
    db.clear_approved()
    await callback.answer("🗑 Ro'yxat tozalandi!", show_alert=True)
    await safe_edit(
        callback.message,
        f"🪪 <b>A'ZOLIK GUVOHNOMASI</b>\n{'─' * 28}\n"
        f"📋 Tasdiqlangan a'zolar ro'yxati: <b>{db.get_approved_count()}</b>\n"
        f"🪪 Berilgan guvohnomalar: <b>{db.get_card_count()}</b>",
        reply_markup=_guv_admin_kb()
    )


@router.message(F.text == "🏠 Asosiy menyu")
async def back_to_main_text(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("Asosiy menyu:", reply_markup=main_menu)


@router.callback_query(F.data == "back_to_main")
async def back_to_main_callback(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    try:
        await callback.message.delete()
    except Exception:
        pass
    await callback.message.answer("Asosiy menyu:", reply_markup=main_menu)
    await callback.answer()


class AdminReplyState(StatesGroup):
    waiting_reply = State()


@router.callback_query(F.data.startswith("reply_user_"))
async def reply_user_btn(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True); return
    user_id = int(callback.data.split("_")[2])
    if db.is_question_answered(user_id):
        await callback.answer("⚠️ Bu savolga allaqachon javob berilgan!", show_alert=True); return
    await state.set_state(AdminReplyState.waiting_reply)
    await state.update_data(reply_target_id=user_id, reply_msg_id=callback.message.message_id)
    await callback.message.answer(
        f"✍️ <b>Foydalanuvchi (ID: {user_id})ga javobingizni yozing:</b>\n\nBekor qilish uchun /admin",
        parse_mode=ParseMode.HTML
    )
    await callback.answer()


@router.message(AdminReplyState.waiting_reply, F.chat.type == "private")
async def admin_send_reply(message: Message, state: FSMContext, bot: Bot):
    if not is_admin(message.from_user.id): return
    data = await state.get_data()
    target_id = data.get("reply_target_id")
    try:
        await bot.send_message(
            target_id,
            f"☎️ <b>Admindan javob keldi:</b>\n\n{message.text}",
            parse_mode=ParseMode.HTML
        )
        await message.answer("✅ Javob yetkazildi!")
        db.mark_question_answered(target_id, message.from_user.id)
        answerer_name = message.from_user.full_name
        for aid in get_all_admin_ids():
            if aid == message.from_user.id: continue
            try:
                await bot.send_message(
                    aid,
                    f"✅ <b>Savolga javob berildi</b>\n\n"
                    f"👤 Javob bergan: {answerer_name}\n"
                    f"🆔 Foydalanuvchi ID: {target_id}",
                    parse_mode=ParseMode.HTML
                )
            except Exception:
                pass
    except Exception as e:
        await message.answer(f"❌ Xatolik: {e}")
    await state.clear()


@router.message(F.reply_to_message, F.chat.type == "private")
async def admin_reply_handler(message: Message, bot: Bot):
    if not is_admin(message.from_user.id): return
    try:
        original_text = message.reply_to_message.text or message.reply_to_message.caption or ""
        user_id = None

        # ID qidirish: ID: 123456 yoki 🆔 123456 formatida
        import re
        # ID: 123456 formati
        match = re.search(r'(?:ID:|🆔)\s*(\d+)', original_text)
        if match:
            user_id = int(match.group(1))

        # Agar user_id topilmasa, reply_to_message.from_user.id dan olish
        if not user_id:
            # Agar reply qilingan xabar foydalanuvchidan bo'lsa
            if message.reply_to_message.from_user:
                user_id = message.reply_to_message.from_user.id
            # Yoki xabar forward qilingan bo'lsa
            elif message.reply_to_message.forward_from:
                user_id = message.reply_to_message.forward_from.id

        if not user_id:
            await message.answer("❌ Bu xabarda user ID si topilmadi.\n\nXabar ichida ID: 123456 bo'lishi kerak, yoki reply qilingan xabarda foydalanuvchi ma'lumoti bo'lishi kerak.")
            return

        if db.is_question_answered(user_id):
            await message.answer("⚠️ Bu savolga allaqachon boshqa admin javob bergan!")
            return
        await bot.send_message(
            user_id,
            f"☎️ <b>Admindan javob keldi:</b>\n\n{message.text}",
            parse_mode=ParseMode.HTML
        )
        await message.answer("✅ Yetkazildi!")
        db.mark_question_answered(user_id, message.from_user.id)
        answerer_name = message.from_user.full_name
        for aid in get_all_admin_ids():
            if aid == message.from_user.id: continue
            try:
                await bot.send_message(
                    aid,
                    f"✅ <b>Savolga javob berildi</b>\n\n"
                    f"👤 Javob bergan: {answerer_name}\n"
                    f"🆔 Foydalanuvchi ID: {user_id}",
                    parse_mode=ParseMode.HTML
                )
            except Exception:
                pass
    except Exception as e:
        await message.answer(f"Xatolik: {e}")


@router.message(F.text == "/admins")
async def show_admins(message: Message):
    # Guruhlarda /admins ishlamaydi
    if message.chat.type in ("group", "supergroup"):
        return
    if message.from_user.id in ADMIN_IDS:
        await message.answer(
            "<b>Adminlar:</b>\n" + "\n".join(f"👤 {a}" for a in ADMIN_IDS),
            parse_mode=ParseMode.HTML
        )
    else:
        await message.answer("Bu buyruq faqat adminlar uchun!")



# =============================================================================
# MODERATSIYA: GURUH HANDLER'LARI
# =============================================================================

async def _mod_send_violation_to_admins(
    bot, chat_id: int, user_id: int, user_name: str,
    reason: str, found: str
):
    """Qoidabuzarlik haqida adminlarga DM yuborish."""
    notif_key = hashlib.md5(
        f"{chat_id}_{user_id}_{time.time()}".encode()
    ).hexdigest()[:8]

    warn_count = mod_get_warns(chat_id, user_id)
    time_str = datetime.now().strftime("%d.%m.%Y %H:%M")

    text = (
        f"⚠️ <b>Guruhda qoidabuzarlik</b>\n"
        f"{'─' * 30}\n"
        f"👤 Foydalanuvchi: {user_name}\n"
        f"🆔 ID: <code>{user_id}</code>\n"
        f"📌 Sabab: {reason}\n"
        f"🔍 Topildi: <code>{found}</code>\n"
        f"⚡ Ogohlantirishlar: {warn_count}/{WARN_LIMIT}\n"
        f"🕐 Vaqt: {time_str}\n\n"
        f"Amalni tanlang:"
    )

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="⚠️ Warn", callback_data=f"mwarn_{chat_id}_{user_id}_{notif_key}"),
            InlineKeyboardButton(text="🔇 Mute 7kun", callback_data=f"mmute_{chat_id}_{user_id}_{notif_key}"),
        ],
        [
            InlineKeyboardButton(text="🚫 Ban", callback_data=f"mban_{chat_id}_{user_id}_{notif_key}"),
            InlineKeyboardButton(text="✅ Rad etish", callback_data=f"mdismiss_{notif_key}"),
        ],
    ])

    sent_msgs = []
    try:
        admins = await bot.get_chat_administrators(chat_id)
        for admin in admins:
            if admin.user.is_bot:
                continue
            try:
                msg = await bot.send_message(
                    admin.user.id, text,
                    reply_markup=kb, parse_mode=ParseMode.HTML
                )
                sent_msgs.append({
                    "chat_id": admin.user.id,
                    "message_id": msg.message_id,
                    "text": text
                })
            except Exception:
                pass
    except Exception:
        pass

    if sent_msgs:
        mod_save_notification(notif_key, sent_msgs)


async def _mod_process_violation(message: Message, reason: str, found: str):
    """Xabarnini o'chirish va adminlarga xabar berish."""
    chat_id = message.chat.id
    user_id = message.from_user.id if message.from_user else 0
    user_name = message.from_user.full_name if message.from_user else f"ID:{user_id}"

    try:
        await message.delete()
    except Exception:
        pass

    await _mod_send_violation_to_admins(
        message.bot, chat_id, user_id, user_name, reason, found
    )


# ── Yangi a'zo xush kelibsiz ──────────────────────────────────────────────

async def welcome_new_member(event: ChatMemberUpdated):
    """Guruhga yangi a'zo qo'shilganda xush kelibsiz."""
    member = event.new_chat_member
    if not member or not member.user or member.user.is_bot:
        return

    chat_id = event.chat.id
    user = member.user
    chat_title = event.chat.title or "guruh"

    # FAQAT guruh va superguruhlarda ishlaydi, kanallarda emas
    if event.chat.type not in ("group", "supergroup"):
        return

    text = (
        f"🌟 Assalomu alaykum, <b>{user.full_name}</b>!\n\n"
        f"Siz <b>{chat_title}</b> guruhiga xush kelibsiz.\n\n"
        f"📋 Guruh qoidalariga rioya qiling:\n"
        f"• Spam va reklama taqiqlanadi\n"
        f"• Havolalar taqiqlanadi\n"
        f"• Bir-biringizni hurmat qiling\n\n"
        f"Savollar uchun adminlarga murojaat qiling."
    )
    try:
        await event.bot.send_message(
            chat_id=chat_id, text=text, parse_mode=ParseMode.HTML
        )
    except Exception as e:
        logging.warning(f"[welcome] {e}")


async def bot_added_to_group(event: ChatMemberUpdated):
    """Bot guruhga qo'shilganda menyusini olib tashlash."""
    # Faqat guruh va superguruhlarda ishlaydi
    if event.chat.type not in ("group", "supergroup"):
        return

    # Bot o'ziga o'zi member bo'lganligini tekshirish
    if event.new_chat_member and event.new_chat_member.user.id == event.bot.id:
        # Bot guruhga qo'shildi, menyuni olib tashlash
        try:
            from aiogram.types import ReplyKeyboardRemove
            await event.bot.send_message(
                chat_id=event.chat.id,
                text="🤖 Bot guruhga qo'shildi! Moderatsiya funksiyalari faollashdi.",
            )
        except Exception as e:
            logging.warning(f"[bot_added] Xabar yuborishda xato: {e}")


# ── Content filter ─────────────────────────────────────────────────────────

@router.message(F.text, F.chat.type.in_({"group", "supergroup"}))
async def mod_filter_text(message: Message):
    """Guruh xabarlarini tekshirish."""
    if not message.from_user:
        return

    # Adminlarni o'tkazib yuborish
    try:
        member = await message.bot.get_chat_member(
            message.chat.id, message.from_user.id
        )
        if member.status in ("administrator", "creator"):
            return
    except Exception:
        pass

    text = message.text or ""

    if CHECK_BANNED_WORDS:
        found = filter_check_banned_word(text)
        if found:
            await _mod_process_violation(message, "Taqiqlangan so'z", found)
            return

    if CHECK_HACKER_WORDS:
        found = filter_check_hacker_word(text)
        if found:
            await _mod_process_violation(message, "Xakerlik mazmuni", found)
            return

    if CHECK_LINKS:
        found = filter_check_link(text)
        if found:
            await _mod_process_violation(message, "Havola/reklama", found)
            return

    if CHECK_SHELL_PATTERNS:
        found = filter_check_shell(text)
        if found:
            await _mod_process_violation(message, "Shell buyruq", found)
            return


@router.message(
    F.document | F.audio | F.video | F.animation,
    F.chat.type.in_({"group", "supergroup"})
)
async def mod_filter_files(message: Message):
    """Guruhda havfli fayllarni bloklash."""
    if not message.from_user or not BLOCK_DANGEROUS_FILES:
        return

    try:
        member = await message.bot.get_chat_member(
            message.chat.id, message.from_user.id
        )
        if member.status in ("administrator", "creator"):
            return
    except Exception:
        pass

    file_obj = message.document or message.audio or message.video or message.animation
    if not file_obj:
        return

    filename = getattr(file_obj, "file_name", "") or ""
    found = filter_check_file_ext(filename)
    if found:
        await _mod_process_violation(message, "Havfli fayl", f".{found}")


# ── Admin buyruqlari (/warn, /mute, /ban ...) ──────────────────────────────

@router.message(F.text.startswith("/warn"))
async def mod_cmd_warn(message: Message):
    if message.from_user.id not in get_all_admin_ids():
        return
    if message.chat.type not in ("group", "supergroup"):
        await message.answer("❌ /warn buyrug'i faqat guruhlarda ishlaydi.")
        return
    if not message.reply_to_message:
        await message.answer("⚠️ Ogohlantirish uchun xabarga javob bering.")
        return

    target = message.reply_to_message.from_user
    if not target:
        return

    try:
        mem = await message.bot.get_chat_member(message.chat.id, target.id)
        if mem.status in ("administrator", "creator"):
            await message.answer("❌ Administratorni ogohlantirish mumkin emas!")
            return
    except Exception:
        pass

    count = mod_add_warn(message.chat.id, target.id)
    days = MUTE_DAYS_PER_WARN * count

    if count >= WARN_LIMIT:
        try:
            await message.bot.ban_chat_member(message.chat.id, target.id)
            mod_reset_warns(message.chat.id, target.id)
            await message.answer(
                f"🚫 <b>{target.full_name}</b> — {WARN_LIMIT} ogohlantirish to'ldi, "
                f"guruhdan chiqarildi!",
                parse_mode=ParseMode.HTML
            )
        except Exception as e:
            await message.answer(f"Xatolik: {e}")
        return

    try:
        until = datetime.now() + timedelta(days=days)
        await message.bot.restrict_chat_member(
            chat_id=message.chat.id, user_id=target.id,
            permissions=ChatPermissions(can_send_messages=False),
            until_date=until
        )
    except Exception:
        pass

    await message.answer(
        f"⚠️ <b>{target.full_name}</b> ogohlantirildi.\n"
        f"Jazo: {days} kun mute\n"
        f"Ogohlantirish: {count}/{WARN_LIMIT}",
        parse_mode=ParseMode.HTML
    )


@router.message(F.text.startswith("/unwarn"))
async def mod_cmd_unwarn(message: Message):
    if message.from_user.id not in get_all_admin_ids():
        return
    # Faqat guruhlarda ishlaydi
    if message.chat.type not in ("group", "supergroup"):
        await message.answer("❌ /unwarn buyrug'i faqat guruhlarda ishlaydi.")
        return
    if not message.reply_to_message:
        await message.answer("Ogohlantirishni olib tashlash uchun xabarga javob bering.")
        return
    target = message.reply_to_message.from_user
    if not target:
        return
    mod_reset_warns(message.chat.id, target.id)
    await message.answer(f"✅ <b>{target.full_name}</b> ning ogohlantirishlari olib tashlandi.", parse_mode=ParseMode.HTML)


@router.message(F.text.startswith("/warns"))
async def mod_cmd_warns(message: Message):
    # Faqat guruhlarda ishlaydi
    if message.chat.type not in ("group", "supergroup"):
        await message.answer("❌ /warns buyrug'i faqat guruhlarda ishlaydi.")
        return
    if not message.reply_to_message:
        await message.answer("Ogohlantirishlarni ko'rish uchun xabarga javob bering.")
        return
    target = message.reply_to_message.from_user
    if not target:
        return
    count = mod_get_warns(message.chat.id, target.id)
    await message.answer(
        f"👤 <b>{target.full_name}</b>\n"
        f"⚠️ Ogohlantirishlar: {count}/{WARN_LIMIT}",
        parse_mode=ParseMode.HTML
    )


@router.message(F.text.startswith("/mute"))
async def mod_cmd_mute(message: Message):
    if message.from_user.id not in get_all_admin_ids():
        return
    # Faqat guruhlarda ishlaydi
    if message.chat.type not in ("group", "supergroup"):
        await message.answer("❌ /mute buyrug'i faqat guruhlarda ishlaydi.")
        return
    if not message.reply_to_message:
        await message.answer("Mute uchun xabarga javob bering. Misol: /mute 60")
        return

    target = message.reply_to_message.from_user
    if not target:
        return

    try:
        mem = await message.bot.get_chat_member(message.chat.id, target.id)
        if mem.status in ("administrator", "creator"):
            await message.answer("❌ Administratorni jazolash mumkin emas!")
            return
    except Exception:
        pass

    # Vaqtni aniqlash
    parts = message.text.split()
    minutes = 60
    if len(parts) > 1 and parts[1].isdigit():
        minutes = int(parts[1])

    try:
        until = datetime.now() + timedelta(minutes=minutes)
        await message.bot.restrict_chat_member(
            chat_id=message.chat.id, user_id=target.id,
            permissions=ChatPermissions(can_send_messages=False),
            until_date=until
        )
        duration = f"{minutes} daqiqa" if minutes < 60 else f"{minutes // 60} soat"
        await message.answer(
            f"🔇 <b>{target.full_name}</b> {duration}ga mute qilindi.",
            parse_mode=ParseMode.HTML
        )
    except Exception as e:
        await message.answer(f"Xatolik: {e}")


@router.message(F.text.startswith("/unmute"))
async def mod_cmd_unmute(message: Message):
    if message.from_user.id not in get_all_admin_ids():
        return
    # Faqat guruhlarda ishlaydi
    if message.chat.type not in ("group", "supergroup"):
        await message.answer("❌ /unmute buyrug'i faqat guruhlarda ishlaydi.")
        return
    if not message.reply_to_message:
        await message.answer("Mute olib tashlash uchun xabarga javob bering.")
        return
    target = message.reply_to_message.from_user
    if not target:
        return
    try:
        await message.bot.restrict_chat_member(
            chat_id=message.chat.id, user_id=target.id,
            permissions=ChatPermissions(
                can_send_messages=True,
                can_send_media_messages=True,
                can_send_other_messages=True,
                can_add_web_page_previews=True
            )
        )
        await message.answer(f"🔊 <b>{target.full_name}</b> mute dan chiqarildi.", parse_mode=ParseMode.HTML)
    except Exception as e:
        await message.answer(f"Xatolik: {e}")


@router.message(F.text.startswith("/ban"))
async def mod_cmd_ban(message: Message):
    if message.from_user.id not in get_all_admin_ids():
        return
    # Guruhda reply bilan ishlaydi
    if message.chat.type in ("group", "supergroup"):
        if not message.reply_to_message:
            await message.answer("Ban uchun xabarga javob bering.")
            return
        target = message.reply_to_message.from_user
        if not target:
            await message.answer("❌ Foydalanuvchi aniqlanmadi.")
            return
        try:
            mem = await message.bot.get_chat_member(message.chat.id, target.id)
            if mem.status in ("administrator", "creator"):
                await message.answer("❌ Administratorni ban qilish mumkin emas!")
                return
        except Exception:
            pass
        try:
            await message.bot.ban_chat_member(message.chat.id, target.id)
            await message.answer(f"🚫 <b>{target.full_name}</b> guruhdan chiqarildi.", parse_mode=ParseMode.HTML)
        except Exception as e:
            await message.answer(f"Xatolik: {e}")
    else:
        # Shaxsiy chatda ID bilan ban qilish
        parts = message.text.split()
        if len(parts) < 2 or not parts[1].isdigit():
            await message.answer("Foydalanish (shaxsiy chat): /ban 123456")
            return
        target_id = int(parts[1])
        # Shaxsiy chatda guruh ID si kerak
        await message.answer("Shaxsiy chatda ban qilish uchun: /ban guruh_id foydalanuvchi_id")


@router.message(F.text.startswith("/unban"))
async def mod_cmd_unban(message: Message):
    if message.from_user.id not in get_all_admin_ids():
        return
    # Guruhlarda ishlaydi
    if message.chat.type not in ("group", "supergroup"):
        await message.answer("❌ /unban buyrug'i faqat guruhlarda ishlaydi.")
        return

    parts = message.text.split()
    if len(parts) < 2 or not parts[1].isdigit():
        await message.answer("Foydalanish: /unban 123456789")
        return
    target_id = int(parts[1])
    try:
        await message.bot.unban_chat_member(message.chat.id, target_id)
        await message.answer(f"✅ ID:<code>{target_id}</code> ban dan chiqarildi.", parse_mode=ParseMode.HTML)
    except Exception as e:
        await message.answer(f"Xatolik: {e}")


@router.message(F.text.startswith("/moders"))
async def mod_cmd_panel(message: Message):
    """Moderatsiya buyruqlari ro'yxati."""
    if message.from_user.id not in get_all_admin_ids():
        return
    # Guruhlarda ham, shaxsiy chatda ham ishlashi mumkin
    text = (
        "🛡️ <b>Moderatsiya buyruqlari</b>\n"
        "──────────────────────────────\n"
        "<b>Xabarga javob qilib:</b>\n"
        "• /warn — ogohlantirish + mute\n"
        "• /unwarn — ogohlantirishlarni olib tashlash\n"
        "• /warns — ogohlantirishlar soni\n"
        "• /mute 60 — 60 daqiqaga mute\n"
        "• /unmute — mute olib tashlash\n"
        "• /ban — guruhdan chiqarish\n\n"
        "<b>ID bilan:</b>\n"
        "• /unban 123456 — ban olib tashlash\n\n"
        f"⚠️ Limit: {WARN_LIMIT} ogohlantirish → auto-ban"
    )
    await message.answer(text, parse_mode=ParseMode.HTML)


# ── Admin DM callback (warn/mute/ban/rad) ──────────────────────────────────

@router.callback_query(F.data.startswith("mwarn_"))
async def mod_cb_warn(callback: CallbackQuery):
    parts = callback.data.split("_")
    if len(parts) < 4:
        await callback.answer("Xatolik", show_alert=True)
        return
    _, chat_id, user_id, notif_key = parts[0], int(parts[1]), int(parts[2]), parts[3]

    entry = mod_get_notification(notif_key)
    if entry and entry.get("handled"):
        await callback.answer("Bu shikoyat allaqachon ko'rib chiqildi.", show_alert=True)
        return

    count = mod_add_warn(chat_id, user_id)
    days = MUTE_DAYS_PER_WARN * count

    try:
        until = datetime.now() + timedelta(days=days)
        await callback.bot.restrict_chat_member(
            chat_id=chat_id, user_id=user_id,
            permissions=ChatPermissions(can_send_messages=False),
            until_date=until
        )
    except Exception:
        pass

    mod_mark_handled(notif_key)
    action_text = f"\n\n✅ Ogohlantirish {count}/{WARN_LIMIT} + {days} kun mute — {callback.from_user.first_name}"

    for msg_data in (entry or {}).get("msgs", []):
        try:
            await callback.bot.edit_message_text(
                chat_id=msg_data["chat_id"],
                message_id=msg_data["message_id"],
                text=msg_data["text"] + action_text,
                parse_mode=ParseMode.HTML
            )
        except Exception:
            pass

    await callback.answer(f"Ogohlantirish berildi! {days} kun mute.")


@router.callback_query(F.data.startswith("mmute_"))
async def mod_cb_mute(callback: CallbackQuery):
    parts = callback.data.split("_")
    if len(parts) < 4:
        await callback.answer("Xatolik", show_alert=True)
        return
    _, chat_id, user_id, notif_key = parts[0], int(parts[1]), int(parts[2]), parts[3]

    entry = mod_get_notification(notif_key)
    if entry and entry.get("handled"):
        await callback.answer("Allaqachon ko'rib chiqildi.", show_alert=True)
        return

    days = MUTE_DAYS_PER_WARN
    try:
        until = datetime.now() + timedelta(days=days)
        await callback.bot.restrict_chat_member(
            chat_id=chat_id, user_id=user_id,
            permissions=ChatPermissions(can_send_messages=False),
            until_date=until
        )
    except Exception:
        pass

    mod_mark_handled(notif_key)
    action_text = f"\n\n✅ {days} kun mute — {callback.from_user.first_name}"
    for msg_data in (entry or {}).get("msgs", []):
        try:
            await callback.bot.edit_message_text(
                chat_id=msg_data["chat_id"],
                message_id=msg_data["message_id"],
                text=msg_data["text"] + action_text,
                parse_mode=ParseMode.HTML
            )
        except Exception:
            pass
    await callback.answer(f"{days} kun mute qo'llandi.")


@router.callback_query(F.data.startswith("mban_"))
async def mod_cb_ban(callback: CallbackQuery):
    parts = callback.data.split("_")
    if len(parts) < 4:
        await callback.answer("Xatolik", show_alert=True)
        return
    _, chat_id, user_id, notif_key = parts[0], int(parts[1]), int(parts[2]), parts[3]

    entry = mod_get_notification(notif_key)
    if entry and entry.get("handled"):
        await callback.answer("Allaqachon ko'rib chiqildi.", show_alert=True)
        return

    try:
        await callback.bot.ban_chat_member(chat_id=chat_id, user_id=user_id)
    except Exception:
        pass

    mod_mark_handled(notif_key)
    action_text = f"\n\n🚫 Ban — {callback.from_user.first_name}"
    for msg_data in (entry or {}).get("msgs", []):
        try:
            await callback.bot.edit_message_text(
                chat_id=msg_data["chat_id"],
                message_id=msg_data["message_id"],
                text=msg_data["text"] + action_text,
                parse_mode=ParseMode.HTML
            )
        except Exception:
            pass
    await callback.answer("Foydalanuvchi ban qilindi.")


@router.callback_query(F.data.startswith("mdismiss_"))
async def mod_cb_dismiss(callback: CallbackQuery):
    notif_key = callback.data.split("_", 1)[1]

    entry = mod_get_notification(notif_key)
    if entry and entry.get("handled"):
        await callback.answer("Allaqachon ko'rib chiqildi.", show_alert=True)
        return

    mod_mark_handled(notif_key)
    action_text = f"\n\n❌ Rad etildi — {callback.from_user.first_name}"
    for msg_data in (entry or {}).get("msgs", []):
        try:
            await callback.bot.edit_message_text(
                chat_id=msg_data["chat_id"],
                message_id=msg_data["message_id"],
                text=msg_data["text"] + action_text,
                parse_mode=ParseMode.HTML
            )
        except Exception:
            pass
    await callback.answer("Shikoyat rad etildi.")


# ── /mreport ────────────────────────────────────────────────────────────────

@router.message(F.text.startswith("/mreport"))
async def mod_cmd_report(message: Message):
    """Guruhda qoidabuzarlik haqida shikoyat."""
    if message.chat.type not in ("group", "supergroup"):
        await message.answer("❌ /mreport buyrug'i faqat guruhlarda ishlaydi.")
        return
    if not message.reply_to_message:
        await message.answer(
            "📢 Shikoyat uchun: <code>/mreport sabab</code>\n"
            "(Qoidabuzar xabariga javob qilib yozing)",
            parse_mode=ParseMode.HTML
        )
        return

    target = message.reply_to_message.from_user
    sender = message.from_user
    if not target or not sender:
        return

    try:
        mem = await message.bot.get_chat_member(message.chat.id, target.id)
        if mem.status in ("administrator", "creator"):
            await message.answer("❌ Administratorga shikoyat qilish mumkin emas!")
            return
    except Exception:
        pass

    parts = message.text.split(maxsplit=1)
    reason = parts[1] if len(parts) > 1 else "Sabab ko'rsatilmagan"

    await message.answer(
        f"✅ Shikoyat yuborildi!\n"
        f"👤 Kim haqida: <b>{target.full_name}</b>\n"
        f"📌 Sabab: {reason}",
        parse_mode=ParseMode.HTML
    )

    notif_key = hashlib.md5(
        f"{message.chat.id}_{target.id}_report_{time.time()}".encode()
    ).hexdigest()[:8]

    report_text = (
        f"📢 <b>SHIKOYAT</b>\n"
        f"{'─' * 30}\n"
        f"🏠 Guruh: {message.chat.title}\n"
        f"📨 Kimdan: {sender.full_name}\n"
        f"👤 Kim haqida: {target.full_name} (<code>{target.id}</code>)\n"
        f"📌 Sabab: {reason}\n"
        f"⚠️ Ogohlantirishlar: {mod_get_warns(message.chat.id, target.id)}/{WARN_LIMIT}\n"
        f"🕐 Vaqt: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    )

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="⚠️ Warn", callback_data=f"mwarn_{message.chat.id}_{target.id}_{notif_key}"),
            InlineKeyboardButton(text="🔇 Mute", callback_data=f"mmute_{message.chat.id}_{target.id}_{notif_key}"),
        ],
        [
            InlineKeyboardButton(text="🚫 Ban", callback_data=f"mban_{message.chat.id}_{target.id}_{notif_key}"),
            InlineKeyboardButton(text="✅ Rad etish", callback_data=f"mdismiss_{notif_key}"),
        ],
    ])

    sent_msgs = []
    try:
        admins = await message.bot.get_chat_administrators(message.chat.id)
        for admin in admins:
            if admin.user.is_bot:
                continue
            try:
                msg = await message.bot.send_message(
                    admin.user.id, report_text,
                    reply_markup=kb, parse_mode=ParseMode.HTML
                )
                sent_msgs.append({
                    "chat_id": admin.user.id,
                    "message_id": msg.message_id,
                    "text": report_text
                })
            except Exception:
                pass
    except Exception:
        pass

    if sent_msgs:
        mod_save_notification(notif_key, sent_msgs)


# ── Oxirgi catch-all handler ────────────────────────────────────────────────
# Guruhda bot faqat chaqirilganda javob beradi, shaxsiy chatda doim

@router.message()
async def other_messages(message: Message):
    # Guruhlarda faqat chaqirilganda javob berish
    if message.chat.type in ("group", "supergroup"):
        # Guruhda faqat bot chaqirilganda javob bersin
        text = message.text or message.caption or ""
        bot_username = BOT_USERNAME.lower()
        # Agar xabarda @botusername bo'lmasa, javob bermaydi
        if f"@{bot_username}" not in text.lower():
            return
        # Bot chaqirilganda, buyruq emas bo'lsa, hech narsa qilmaydi
        return

    # Shaxsiy chatda menyu ko'rsatish
    try:
        await message.answer("Iltimos, menyudan biror buyruqni tanlang.", reply_markup=main_menu)
    except Exception as e:
        logging.warning(f"[other_messages] Xabar yuborib bo'lmadi (chat={message.chat.id}): {e}")




# =============================================================================
# ISHGA TUSHIRISH
# =============================================================================
async def main():
    bot = Bot(token=TOKEN)
    dp = Dispatcher()
    dp.include_router(router)

    # ── ANTI-FLOOD MIDDLEWARE ─────────────────────────────────────────────────
    dp.message.middleware(RateLimitMiddleware())

    # ── YANGI A'ZO SALOMLASH ──────────────────────────────────────────────────
    dp.chat_member.register(welcome_new_member, ChatMemberUpdatedFilter(JOIN_TRANSITION))

    # ── BOT GURUHGA QO'SHILGANDA MENYUNI OLIB TASHLASH ──────────────────────────
    dp.my_chat_member.register(bot_added_to_group, ChatMemberUpdatedFilter(JOIN_TRANSITION))

    # ── GLOBAL ERROR HANDLER ─────────────────────────────────────────────────
    @dp.errors()
    async def global_error_handler(event, exception: Exception):
        if isinstance(exception, TelegramForbiddenError):
            # Foydalanuvchi botni bloklagan — jim o'tkazib yuboramiz
            logging.warning(f"[BLOCKED] Foydalanuvchi botni bloklagan: {exception}")
            return True
        if isinstance(exception, TelegramBadRequest):
            msg = str(exception)
            if "query is too old" in msg or "query ID is invalid" in msg:
                # Bot o'chirilgan paytda bosilgan eski tugma — normal holat
                logging.warning(f"[OLD_QUERY] Eski callback, e'tiborsiz: {exception}")
                return True
        # Boshqa xatolar — odatdagidek log yoziladi
        logging.exception(f"[ERROR] Kutilmagan xato: {exception}")
        return False

    print(f"Bot ishga tushdi! Adminlar: {ADMIN_IDS}")
    print(f"Majburiy kanallar: {REQUIRED_CHANNELS or 'O`chirilgan'}")
    print(f"Moderatsiya: warn_limit={WARN_LIMIT}, rate_limit={RATE_LIMIT_MAX}/{RATE_LIMIT_WINDOW}s")
    await dp.start_polling(bot, allowed_updates=dp.resolve_used_update_types())


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, stream=sys.stdout)
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("Bot to'xtatildi")